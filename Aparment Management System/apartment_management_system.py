import tkinter as tk
from tkinter import *
from tkinter import ttk, messagebox
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

excel_file = "apartment_management_excel.xlsx"

# Openpyxl
excel_connection = Workbook()
excel_connection = load_workbook(excel_file)
main_sheet = excel_connection['main']
apart_sheet = excel_connection['apartment']
tenant_main_sheet = excel_connection['tenant']
payment_main_sheet = excel_connection['payment']
payment_record_sheet = excel_connection['payment records']


def log_in_interface():
    global attempts
    log_in_window = tk.Tk()
    log_in_window.title("Log In Page")
    log_in_window.geometry("340x385")
    log_in_window.resizable(0, 0)
    attempts = 0
    
    # Theme
    style = ttk.Style(log_in_window)
    log_in_window.tk.call("source", 'theme_forest-dark.tcl')
    style.theme_use("forest-dark")

    # Functions
    def show_passw():
        if passw_entry.get().strip() == "PASSWORD" or passw_entry.get().strip() == "":
            passw_entry['show'] = ""
        else:
            if passw_var.get() == 1:
                if passw_entry['show'] == "":
                    passw_entry['show'] = "•"
                elif passw_entry['show'] == "•":
                    passw_entry['show'] = ""
            else:
                passw_entry['show'] = "•"


    def log_in():
        username = "1"
        password = "1"
        correct_username = False
        correct_password = False
        entry_username = usern_entry.get().strip()
        entry_password = passw_entry.get()
        if entry_username == username:
            correct_username = True
        if entry_password == password:
            correct_password = True
        if correct_username and correct_password:
            log_in_window.destroy()
            main_interface()
        else:
            passw_out
            messagebox.showerror("ERROR", "Invalid username or password.")
            log_in_attempt(1)


    def log_in_attempt(num):
        global attempts
        attempts += num
        if attempts == 5:
            log_in_window.destroy()
            messagebox.showinfo("ERROR", "You have try so many times. Try again later")

    
    def cancel_login():
        log_in_window.destroy()


    def usern_in(event):
        user_name = usern_entry.get().strip()
        if user_name == "" or user_name == "USERNAME":
            usern_entry.delete(0, END)
    def usern_out(event):
        user_name = usern_entry.get().strip()
        if user_name == "":
            usern_entry.insert(0, "USERNAME")

    def passw_in(event):
        passw = passw_entry.get().strip()
        if passw == "" or passw == "PASSWORD":
            passw_entry.delete(0, END)
            if passw_var.get() == 1:
                passw_entry['show'] = ""
            else:
                passw_entry['show'] = "•"
    def passw_out(event):
        passw = passw_entry.get().strip()
        if passw == "":
            passw_entry.insert(0, "PASSWORD")
            passw_entry['show'] = ""
            passw_var = 0


    # Widgets
    main_frame = ttk.Frame(log_in_window)
    # Welcome Greeting
    greeting_label = Label(main_frame, text="WELCOME", font=("Arial Black", 35, "bold",), fg="#217346")
    # Username
    usern_entry = ttk.Entry(main_frame, width=40, font=("", 13))
    usern_entry.insert(0,"USERNAME")
    usern_entry.bind("<FocusIn>", usern_in)
    usern_entry.bind("<FocusOut>", usern_out)
    # Password
    passw_entry = ttk.Entry(main_frame, show="", width=40, font=("", 13))
    passw_entry.insert(0,"PASSWORD")
    passw_entry.bind("<FocusIn>", passw_in)
    passw_entry.bind("<FocusOut>", passw_out)
    passw_var = IntVar()
    show_passw_btn = ttk.Checkbutton(main_frame, text="SHOW PASSWORD", variable=passw_var, command=show_passw)
    # Login btn
    log_in_btn = Button(main_frame, text="LOG IN", relief='flat', font=("", 10), bg='#217346', command=log_in)
    cancel_btn = ttk.Button(main_frame, text="CANCEL", command=cancel_login)


    # Geometry Manager
    main_frame.pack(expand=True, fill=BOTH)
    # Welcome Greeting
    greeting_label.pack(pady=(20, 20))
    # Username
    usern_entry.pack(padx=40, pady=(0, 15), ipady=5)
    # Password
    passw_entry.pack(padx=40, pady=(0, 5), ipady=5)
    show_passw_btn.pack(padx=45, pady=(0, 20), anchor="nw")
    # Login btn
    log_in_btn.pack(fill=X, ipady=4, padx=38, pady=(0, 7))
    cancel_btn.pack(fill=X, ipady=2, padx=38)
    log_in_window.mainloop()


def main_interface():
    main_interface_root = tk.Tk()
    main_interface_root.resizable(0, 0)
    main_interface_root.title("Apartment Management System")

    # Theme
    style = ttk.Style(main_interface_root)
    main_interface_root.tk.call("source", 'theme_forest-dark.tcl')
    style.theme_use("forest-dark")

    # Functions
    def back_to_login():
        want_to_leave = messagebox.askquestion('Are you sure?', 'Are you sure you want to leave?')
        if want_to_leave == "yes":
            main_interface_root.destroy()


    def search_unit():
        def exit_apart_search():
            apart_search_tree_tl.destroy()
            
        search_unit_number = search_entry.get().strip()
        if search_unit_number != "" and search_unit_number != "Unit #":
            search_unit_data_found = False
            for search_every_cell_apart in range(2, apart_sheet.max_row+1):
                if  search_unit_number == apart_sheet['A'+str(search_every_cell_apart)].value:
                    search_number_row = search_every_cell_apart
                    search_unit_data_found = True 
                    break
            if search_unit_data_found:
                apart_search_tree_tl = tk.Toplevel()
                apart_search_tree_tl.resizable(0,0)
                apart_search_tree_tl.title("Apartment Details")

                apart_search_tree_main_frame = Frame(apart_search_tree_tl)
                apart_search_tree_main_frame.pack(expand=TRUE, fill=BOTH, padx=10, pady=10)
                cols = ["1", "2"]
                apart_search_tree = ttk.Treeview(apart_search_tree_main_frame, show='headings', columns=cols, height=18)
                apart_search_tree.column("1", anchor=W, width=150)
                apart_search_tree.heading("1", text="APARTMENT's", anchor=E)
                apart_search_tree.column("2", anchor=W, width=150)
                apart_search_tree.heading("2", text="  DATA", anchor=W)
                list_of_data = [] 
                unit_heading = apart_sheet[1]
                unit_data = apart_sheet[search_number_row]
                ind = 0
                for every_data in unit_data:
                    heading = unit_heading[ind]
                    list_of_data.append([heading.value, every_data.value])
                    ind += 1
                for each_data in list_of_data:
                    apart_search_tree.insert("", END, values=each_data)
                apart_search_tree.pack(expand=TRUE, fill=Y)
                
                has_tenant = False
                for every_tenant in range(2, tenant_main_sheet.max_row+1):
                    if tenant_main_sheet["A"+str(every_tenant)].value == search_unit_number:
                        has_tenant = True
                        tenant_row = every_tenant
                        break
                if has_tenant:
                    cols = ["1", "2"]
                    tenant_search_tree = ttk.Treeview(apart_search_tree_main_frame, show='headings', columns=cols, height=6)
                    tenant_search_tree.column("1", anchor=W, width=150)
                    tenant_search_tree.heading("1", text="TENANT's", anchor=E)
                    tenant_search_tree.column("2", anchor=W, width=150)
                    tenant_search_tree.heading("2", text="  DATA", anchor=W)
                    list_of_data = [] 
                    unit_heading = tenant_main_sheet[1]
                    unit_data = tenant_main_sheet[search_number_row]
                    ind = 0
                    for every_data in unit_data:
                        heading = unit_heading[ind]
                        list_of_data.append([heading.value, every_data.value])
                        ind += 1
                    for each_data in list_of_data:
                        tenant_search_tree.insert("", END, values=each_data)
                    tenant_search_tree.pack(expand=TRUE, fill=Y, pady=(5, 0))
                exit_btn = ttk.Button(apart_search_tree_main_frame, text="EXIT", command=exit_apart_search)
                exit_btn.pack(fill=X, pady=(5, 0))
            else:
                search_entry.delete(0, END)
                messagebox.showerror("ERROR", f"There is no existing data for \"{search_unit_number}\".")
        else:
            messagebox.showerror('ERROR', "Must provide a Unit Number.")


    def manage_apartment():
        main_interface_root.destroy()
        apartment()


    def manage_tenant():
        main_interface_root.destroy()
        tenant()


    def manage_payment():
        main_interface_root.destroy()
        payment()

    # Widgets
    main_frame = Frame(main_interface_root)
    top_frame_title = Frame(main_frame)
    title = Label(top_frame_title, text="Aparment Management System", font=("Arial Black", 40),
                  fg="#217346")

    middle_frame = Frame(main_frame)
    left_frame = Frame(middle_frame)
    right_frame = Frame(middle_frame)
    excel_items = list(main_sheet.values)
    tree = ttk.Treeview(left_frame, show='headings', columns=excel_items[0], height=15)
    scroll_bar_y = ttk.Scrollbar(left_frame, orient='vertical', command=tree.yview)
    tree.configure(yscrollcommand=scroll_bar_y.set)
    for heading in excel_items[0]:
        if heading ==  "UNIT":
            tree.column(heading, anchor=CENTER, width=40)
            tree.heading(heading, text=heading)
        elif heading ==  "STATUS" or heading == 'PRICE':
            tree.column(heading, anchor=CENTER, width=100)
            tree.heading(heading, text=heading)
        else:
            tree.column(heading, anchor=CENTER)
            tree.heading(heading, text=heading)
    for value in excel_items[1:]:
        tree.insert('', END, values=value)

    search_frame = ttk.LabelFrame(right_frame, text="Search")
    search_entry = ttk.Entry(search_frame, font=("", 13, "bold"))
    search_entry.insert(0,"Unit Number")
    def s_i(event):
        if search_entry.get().strip() == "" or \
            search_entry.get().strip() == "Unit Number":
            search_entry.delete(0, END)
    def s_o(event):
        if search_entry.get().strip() == "":
            search_entry.insert(0, "Unit Number")
    search_entry.bind("<FocusIn>", s_i)
    search_entry.bind("<FocusOut>", s_o)
    search_btn = Button(search_frame, text="SEARCH", relief='flat', font=("", 10, "bold"), bg='#217346', command=search_unit)
    management_frame = ttk.LabelFrame(right_frame, text='Management')
    apart_man_btn = Button(management_frame, text="MANAGE APARMENTS", relief='flat', font=("", 12, "bold"), bg='#217346', command=manage_apartment)
    tennat_man_btn = Button(management_frame, text="MANAGE TENANTS", relief='flat', font=("", 12, "bold"), bg='#217346', command=manage_tenant)
    paym_btn = Button(management_frame, text="MANAGE PAYMENTS", relief='flat', font=("", 12, "bold"), bg='#217346', command=manage_payment)
    exit_btn = ttk.Button(right_frame, text="EXIT", command=back_to_login)

    # Geometry
    main_frame.pack(expand=TRUE, padx=20, pady=(10, 20))
    top_frame_title.pack(pady=(0, 10), fill=X)
    title.pack(expand=TRUE)

    middle_frame.pack(expand=TRUE, fill=BOTH)
    left_frame.pack(side=LEFT, expand=TRUE, fill=BOTH, pady=(5, 0))
    tree.pack(expand=TRUE, fill=BOTH, side=LEFT)
    scroll_bar_y.pack(fill=Y, side=RIGHT)

    right_frame.pack(side=RIGHT, fill=BOTH, padx=(10, 0), expand=TRUE)

    search_frame.pack(fill=BOTH, pady=(0, 5))
    search_entry.pack(fill=BOTH, padx=5, pady=5)
    search_btn.pack(fill=X, padx=5, pady=(0, 5))

    management_frame.pack(expand=TRUE, fill=BOTH, pady=(0, 10))
    apart_man_btn.pack(expand=TRUE, fill=BOTH, padx=5, pady=10)
    tennat_man_btn.pack(expand=TRUE, fill=BOTH, padx=5, pady=(0, 10))
    paym_btn.pack(expand=TRUE, fill=BOTH, padx=5, pady=(0, 5))
    exit_btn.pack(ipady=3, fill=BOTH)

    main_interface_root.mainloop()


def apartment():
    manage_apart = tk.Tk()
    manage_apart.geometry("1300x723")
    manage_apart.resizable(0, 0)
    manage_apart.title("Apartment Management")

    # Theme
    style = ttk.Style(manage_apart)
    manage_apart.tk.call("source", 'theme_forest-dark.tcl')
    style.theme_use("forest-dark")

    def is_unit_occupied_func(unit_num):
        is_unit_occupied = False
        occupied_units = list(tenant_main_sheet["A"])
        for every_unit_num in occupied_units[1:]:
            if every_unit_num.value == unit_num:
                is_unit_occupied = True
                break
        return is_unit_occupied


    def apart_to_update(unit_number, verify):
        verify.destroy()
        number_row = 0
        for every_cell_apart in range(2, apart_sheet.max_row+1):
            if  unit_number == apart_sheet['A'+str(every_cell_apart)].value:
                number_row = every_cell_apart
                break
        apart_update_tl = tk.Tk()
        apart_update_tl.resizable(0, 0)
        apart_update_tl.title("Update Page")
        style = ttk.Style(apart_update_tl)
        apart_update_tl.tk.call("source", 'theme_forest-dark.tcl')
        style.theme_use("forest-dark")
        update_tl_apart_l = []
        update_tl_apart_sheet_l = list(apart_sheet[number_row])
        for update_tl_apart_every_item in update_tl_apart_sheet_l:
            update_tl_apart_l.append(update_tl_apart_every_item.value)

        def apart_have_found():
            apart_update_tl_unit_entry.insert(0, update_tl_apart_l[0])
            apart_update_tl_unit_var.set(1)


        def apart_update_tl_update_func():
            apartment_unit_list = list(apart_sheet["A"])
            apartment_unit_list_value = []
            for every_number in apartment_unit_list:
                apartment_unit_list_value.append(every_number.value)
            if apart_update_tl_unit_entry.get().strip() != unit_number and \
                apart_update_tl_unit_entry.get().strip() in apartment_unit_list_value:
                messagebox.showerror("ERROR", f"Unit {apart_update_tl_unit_entry.get().strip()} already exist.")
            else:
                is_unit_has_tenant = False
                if apart_update_tl_status_var_s.get().strip() == "Vacant":
                    _is_unit_occupied = is_unit_occupied_func(unit_number)
                    if _is_unit_occupied:
                        is_unit_has_tenant = True
                if is_unit_has_tenant:
                    messagebox.showerror("ERROR", f"Can't set apartment status to {apart_update_tl_status_var_s.get().strip()}, unit {unit_number} is currently has tenant in record.\n\nDelete the tenant in unit {unit_number} in Tenant Management first.")
                    apart_update_tl_status_var_s.set("")
                    apart_update_tl.destroy()
                    tenant()
                elif is_unit_has_tenant == False:
                    if apart_update_tl_unit_entry.get().strip() == "":
                        updated_value_unit_number = apart_sheet[f"A{number_row}"].value
                    else:
                        updated_value_unit_number = apart_update_tl_unit_entry.get().strip()

                    if apart_update_tl_rate_entry.get().strip() == "":
                        updated_value_rate = apart_sheet[f"B{number_row}"].value
                    else:
                        updated_value_rate = apart_update_tl_rate_entry.get().strip()
                        
                    if apart_update_tl_status_var_s.get().strip() == "":
                        updated_value_status = apart_sheet[f"C{number_row}"].value
                    else:
                        updated_value_status = apart_update_tl_status_var_s.get().strip()

                    if apart_update_tl_size_entry.get().strip() == "":
                        updated_value_size = apart_sheet[f"D{number_row}"].value
                    else:
                        updated_value_size = apart_update_tl_size_entry.get().strip()

                    if apart_update_tl_spin.get().strip() == "":
                        updated_value_floor = apart_sheet[f"E{number_row}"].value
                    else:
                        updated_value_floor = apart_update_tl_spin.get().strip()

                    if apart_update_tl_bedrooms_spin.get().strip() == "":
                        updated_value_bedroom = apart_sheet[f"F{number_row}"].value
                    else:
                        updated_value_bedroom = apart_update_tl_bedrooms_spin.get().strip()

                    if apart_update_tl_bathroom_spin.get().strip() == "":
                        updated_value_bathroom = apart_sheet[f"G{number_row}"].value
                    else:
                        updated_value_bathroom = apart_update_tl_bathroom_spin.get().strip()

                    if apart_update_tl_kitchen_var_s.get().strip() == "":
                        updated_value_kitchen = apart_sheet[f"H{number_row}"].value
                    else:
                        updated_value_kitchen = apart_update_tl_kitchen_var_s.get().strip()

                    if update_tl_apart_livingroom_var_s.get().strip() == "":
                        updated_value_living = apart_sheet[f"I{number_row}"].value
                    else:
                        updated_value_living = update_tl_apart_livingroom_var_s.get().strip()

                    if update_tl_apart_dining_var_s.get().strip() == "":
                        updated_value_dining = apart_sheet[f"J{number_row}"].value
                    else:
                        updated_value_dining = update_tl_apart_dining_var_s.get().strip()

                    if update_tl_apart_balcony_var_s.get().strip() == "":
                        updated_value_balcony = apart_sheet[f"K{number_row}"].value
                    else:
                        updated_value_balcony = update_tl_apart_balcony_var_s.get().strip()

                    if update_tl_apart_storage_var_s.get().strip() == "":
                        updated_value_storage = apart_sheet[f"L{number_row}"].value
                    else:
                        updated_value_storage = update_tl_apart_storage_var_s.get().strip()

                    if update_tl_apart_laundry_var_s.get().strip() == "":
                        updated_value_laundry = apart_sheet[f"M{number_row}"].value
                    else:
                        updated_value_laundry = update_tl_apart_laundry_var_s.get().strip()

                    if update_tl_apart_water_var_s.get().strip() == "":
                        updated_value_water = apart_sheet[f"N{number_row}"].value
                    else:
                        updated_value_water = update_tl_apart_water_var_s.get().strip()
                    if apart_update_tl_water_rate_frame.get().strip() == "Water" or apart_update_tl_water_rate_frame.get().strip() == "":
                        if updated_value_water == "NONE":
                            if apart_sheet[f"O{number_row}"].value != "NONE":
                                updated_value_water_rate = "NONE"
                            else:
                                updated_value_water_rate = apart_sheet[f"O{number_row}"].value
                        else:
                            updated_value_water_rate = apart_sheet[f"O{number_row}"].value
                    else:
                        updated_value_water_rate = apart_update_tl_water_rate_frame.get().strip()

                    if update_tl_apart_electricity_var_s.get().strip() == "":
                        updated_value_electricity = apart_sheet[f"P{number_row}"].value
                    else:
                        updated_value_electricity = update_tl_apart_electricity_var_s.get().strip()
                    if apart_update_tl_elec_rate_frame.get().strip() == "Electricity" or apart_update_tl_elec_rate_frame.get().strip() == "":
                        if updated_value_electricity == "NONE":
                            if apart_sheet[f"Q{number_row}"].value != "NONE":
                                updated_value_elec_rate = "NONE"
                            else:
                                updated_value_elec_rate = apart_sheet[f"Q{number_row}"].value
                        else:
                            updated_value_elec_rate = apart_sheet[f"Q{number_row}"].value
                    else:
                        updated_value_elec_rate = apart_update_tl_elec_rate_frame.get().strip()

                    if update_tl_apart_internet_var_s.get().strip() == "":
                        updated_value_internet = apart_sheet[f"R{number_row}"].value
                    else:
                        updated_value_internet = update_tl_apart_internet_var_s.get().strip()
                    if apart_update_tl_inter_rate_frame.get().strip() == "" or apart_update_tl_inter_rate_frame.get().strip() == "Internet":
                        if updated_value_internet == "NONE":
                            if apart_sheet[f"S{number_row}"].value != "NONE":
                                updated_value_internet_rate = "NONE"
                            else:
                                updated_value_internet_rate = apart_sheet[f"S{number_row}"].value
                        else:
                            updated_value_internet_rate = apart_sheet[f"S{number_row}"].value
                    else:
                        updated_value_internet_rate = apart_update_tl_inter_rate_frame.get().strip()
                    
                    water_rate_set = True
                    elec_rate_set = True
                    inter_rate_set = True
                    if updated_value_water == "INCLUDED":
                        if updated_value_water_rate == "NONE" or updated_value_water_rate == "Water" or updated_value_water_rate == "":
                            water_rate_set = False
                    if  updated_value_electricity == "INCLUDED":
                        if updated_value_elec_rate == "NONE" or updated_value_elec_rate == "Electricity" or updated_value_elec_rate == "":
                            elec_rate_set = False
                    if  updated_value_internet == "INCLUDED":
                        if updated_value_internet_rate == "NONE" or updated_value_internet_rate == "Internet" or updated_value_internet_rate == "":
                            inter_rate_set = False
                    if water_rate_set and elec_rate_set and inter_rate_set:
                        update_apartment_answer = messagebox.askquestion("Verifying", f"YOUR ABOUT TO UPDATE THE DATA FOR \"{unit_number}\".\n\nNote: THIS WILL MAKE CHANGE TO THE FILE PERMANENTLY\n\nCLICK \"YES\" TO CONTINUE.")
                        if update_apartment_answer == "yes":
                            for every_apart in range(2, main_sheet.max_row+1):
                                if unit_number == main_sheet['A'+str(every_apart)].value:
                                    unit_main_row = every_apart
                                    break
                            main_sheet[f"A{unit_main_row}"] = updated_value_unit_number
                            main_sheet[f"B{unit_main_row}"] = updated_value_status
                            main_sheet[f"D{unit_main_row}"] = updated_value_rate

                            apart_sheet[f"A{number_row}"] = updated_value_unit_number
                            apart_sheet[f"B{number_row}"] = updated_value_rate
                            apart_sheet[f"C{number_row}"] = updated_value_status
                            apart_sheet[f"D{number_row}"] = updated_value_size
                            apart_sheet[f"E{number_row}"] = updated_value_floor
                            apart_sheet[f"F{number_row}"] = updated_value_bedroom
                            apart_sheet[f"G{number_row}"] = updated_value_bathroom
                            apart_sheet[f"H{number_row}"] = updated_value_kitchen
                            apart_sheet[f"I{number_row}"] = updated_value_living
                            apart_sheet[f"J{number_row}"] = updated_value_dining
                            apart_sheet[f"K{number_row}"] = updated_value_balcony
                            apart_sheet[f"L{number_row}"] = updated_value_storage
                            apart_sheet[f"M{number_row}"] = updated_value_laundry
                            apart_sheet[f"N{number_row}"] = updated_value_water
                            apart_sheet[f"O{number_row}"] = updated_value_water_rate
                            apart_sheet[f"P{number_row}"] = updated_value_electricity
                            apart_sheet[f"Q{number_row}"] = updated_value_elec_rate
                            apart_sheet[f"R{number_row}"] = updated_value_internet
                            apart_sheet[f"S{number_row}"] = updated_value_internet_rate
                            t_found = False
                            t_row = 0
                            for every_t in range(2, tenant_main_sheet.max_row+1):
                                if tenant_main_sheet[f"A{every_t}"].value == unit_number:
                                    t_found = True
                                    t_row = every_t
                                    break
                            if t_found:
                                tenant_main_sheet[f"A{t_row}"] = updated_value_unit_number
                                excel_connection.save(excel_file)
                            p_found = False
                            p_row = 0
                            for every_d in range(2, payment_main_sheet.max_row+1):
                                if payment_main_sheet[f"A{every_d}"].value == unit_number:
                                    p_found = True
                                    p_row = every_d
                                    break
                            if p_found:
                                payment_main_sheet[f"A{p_row}"] = updated_value_unit_number
                                excel_connection.save(excel_file)
                            record_row = 0
                            record_found = False
                            for every_record in range(1, payment_record_sheet.max_row+1):
                                if unit_number == payment_record_sheet[f"A{every_record}"].value:
                                    record_row = every_record
                                    record_found = True
                                    break
                            if record_found:
                                payment_record_sheet["A"+str(record_row)] = updated_value_unit_number
                                excel_connection.save(excel_file)
                            excel_connection.save(excel_file)
                            if updated_value_status == "Occupied" and t_found == False:
                                messagebox.showinfo("Add tenant", f"Add the New Tenant to the New Apartment.")
                                apart_update_tl.destroy()
                                tenant()
                            else:
                                apart_update_tl.destroy()
                                apartment()
                    else:
                        messagebox.showerror("ERROR", "Must set Utilities Rate.")


        def apart_update_tl_exit_func():
            apart_update_tl.destroy()
            apartment()


        apart_update_tl_frame = Frame(apart_update_tl)
        apart_update_tl_frame.pack(padx=10, pady=(0,10) , expand=TRUE, fill=BOTH)

        apart_update_tl_top_frame = Frame(apart_update_tl_frame)
        apart_update_tl_top_frame.pack()

        apart_update_tl_title = Label(apart_update_tl_top_frame, text=f"UPDATE UNIT \"{update_tl_apart_l[0]}\"", font=("Arial Black", 40),
                        fg="#217346")
        apart_update_tl_title.pack()


        apart_update_tl_bottom_frame = ttk.LabelFrame(apart_update_tl_frame, text="UPDATE UNIT")
        apart_update_tl_bottom_frame.pack()
        
        apart_update_tl_bottom_frame_ente = Frame(apart_update_tl_bottom_frame)
        apart_update_tl_bottom_frame_ente.pack()
        apart_update_tl_rigth_frame = Frame(apart_update_tl_bottom_frame_ente)
        apart_update_tl_rigth_frame.pack(side=LEFT, padx=10)


        def get_existing_unit_number():
            if apart_update_tl_unit_var.get() == 0:
                apart_update_tl_unit_entry.delete(0, END)
            else:
                apart_update_tl_unit_entry.delete(0, END)
                apart_update_tl_unit_entry.insert(0, update_tl_apart_l[0])
        apart_update_tl_unit_frame = Frame(apart_update_tl_rigth_frame)
        apart_update_tl_unit_label = ttk.Label(apart_update_tl_unit_frame, text="Unit #:", font=("", 13, "bold"))
        apart_update_tl_unit_entry = ttk.Entry(apart_update_tl_unit_frame, width=28)
        apart_update_tl_unit_var = IntVar()
        apart_update_tl_unit_check = ttk.Checkbutton(apart_update_tl_unit_frame, text="Retrieve old data", variable=apart_update_tl_unit_var, command=get_existing_unit_number)
        apart_update_tl_unit_frame.pack(expand=TRUE, fill=BOTH)
        apart_update_tl_unit_label.grid(row=0, column=0, sticky=W, padx=(10,50))
        apart_update_tl_unit_entry.grid(row=0, column=1)
        apart_update_tl_unit_check.grid(row=1, column=1, sticky=W, pady=5, padx=10)


        def get_existing_unit_status():
            if apart_update_tl_status_var.get() == 0:
                apart_update_tl_status_combo.delete(0, END)
            else:
                apart_update_tl_status_combo.delete(0, END)
                apart_update_tl_status_combo.insert(0, update_tl_apart_l[2])
        apart_update_tl_status_frame = Frame(apart_update_tl_rigth_frame)
        apart_update_tl_status_label = ttk.Label(apart_update_tl_status_frame, text="Unit status:", font=("", 13, "bold"))
        apart_update_tl_status = [
            "Under Construction",
            "Vacant",
            "Occupied"
        ]
        apart_update_tl_status_var_s = StringVar()
        apart_update_tl_status_combo = ttk.Combobox(apart_update_tl_status_frame, textvariable=apart_update_tl_status_var_s, values=apart_update_tl_status, width=25)
        apart_update_tl_status_var = IntVar()
        apart_update_tl_status_check = ttk.Checkbutton(apart_update_tl_status_frame, text="Retrieve old data", variable=apart_update_tl_status_var, command=get_existing_unit_status)
        apart_update_tl_status_frame.pack(expand=TRUE, fill=BOTH)
        apart_update_tl_status_label.grid(row=0, column=0, padx=(10,10))
        apart_update_tl_status_combo.grid(row=0, column=1)
        apart_update_tl_status_check.grid(row=1, column=1, sticky=W, pady=5, padx=10)


        def get_existing_size():
            if apart_update_tl_size_var.get() == 0:
                apart_update_tl_size_entry.delete(0, END)
            else:
                apart_update_tl_size_entry.delete(0, END)
                apart_update_tl_size_entry.insert(0, update_tl_apart_l[3])
        apart_update_tl_size_frame = Frame(apart_update_tl_rigth_frame)
        apart_update_tl_size_label = ttk.Label(apart_update_tl_size_frame, text="Size (m^2):", font=("", 13, "bold"))
        apart_update_tl_size_entry = ttk.Entry(apart_update_tl_size_frame, width=28)
        apart_update_tl_size_var = IntVar()
        apart_update_tl_size_check = ttk.Checkbutton(apart_update_tl_size_frame, text="Retrieve old data", variable=apart_update_tl_size_var, command=get_existing_size)
        apart_update_tl_size_frame.pack(expand=TRUE, fill=BOTH)
        apart_update_tl_size_label.grid(row=0, column=0, padx=(10,12))
        apart_update_tl_size_entry.grid(row=0, column=1)
        apart_update_tl_size_check.grid(row=1, column=1, sticky=W, pady=5, padx=10)


        def get_existing_rate():
            if apart_update_tl_rate_var.get() == 0:
                apart_update_tl_rate_entry.delete(0, END)
            else:
                apart_update_tl_rate_entry.delete(0, END)
                apart_update_tl_rate_entry.insert(0, update_tl_apart_l[1])
        apart_update_tl_rate_frame = Frame(apart_update_tl_rigth_frame)
        apart_update_tl_rate_label = ttk.Label(apart_update_tl_rate_frame, text="Rate:", font=("", 13, "bold"))
        apart_update_tl_rate_entry = ttk.Entry(apart_update_tl_rate_frame, width=28)
        apart_update_tl_rate_var = IntVar()
        apart_update_tl_rate_check = ttk.Checkbutton(apart_update_tl_rate_frame, text="Retrieve old data", variable=apart_update_tl_rate_var, command=get_existing_rate)
        apart_update_tl_rate_frame.pack(expand=TRUE, fill=BOTH)
        apart_update_tl_rate_label.grid(row=0, column=0, padx=(10,60))
        apart_update_tl_rate_entry.grid(row=0, column=1)
        apart_update_tl_rate_check.grid(row=1, column=1, sticky=W, pady=5, padx=10)


        def get_existing_floor():
            if apart_update_tl_floor_var.get() == 0:
                apart_update_tl_spin.delete(0, END)
            else:
                apart_update_tl_spin.delete(0, END)
                apart_update_tl_spin.insert(0, update_tl_apart_l[4])
        apart_update_tl_floor_frame = Frame(apart_update_tl_rigth_frame)
        apart_update_tl_floor_label = ttk.Label(apart_update_tl_floor_frame, text="Floors:", font=("", 13, "bold"))
        apart_update_tl_spin = ttk.Spinbox(apart_update_tl_floor_frame, from_=1, to=100, width=21)
        apart_update_tl_floor_var = IntVar()
        apart_update_tl_floor_check = ttk.Checkbutton(apart_update_tl_floor_frame, text="Retrieve old data", variable=apart_update_tl_floor_var, command=get_existing_floor)
        apart_update_tl_floor_frame.pack(expand=TRUE, fill=BOTH)
        apart_update_tl_floor_label.grid(row=0, column=0, padx=(10,45))
        apart_update_tl_spin.grid(row=0, column=1)
        apart_update_tl_floor_check.grid(row=1, column=1, sticky=W, pady=5, padx=10)


        def get_existing_bedrooms():
            if apart_update_tl_bedrooms_var.get() == 0:
                apart_update_tl_bedrooms_spin.delete(0, END)
            else:
                apart_update_tl_bedrooms_spin.delete(0, END)
                apart_update_tl_bedrooms_spin.insert(0, update_tl_apart_l[5])
        apart_update_tl_bedrooms_frame = Frame(apart_update_tl_rigth_frame)
        apart_update_tl_bedrooms_label = ttk.Label(apart_update_tl_bedrooms_frame, text="Bedrooms:", font=("", 13, "bold"))
        apart_update_tl_bedrooms_spin = ttk.Spinbox(apart_update_tl_bedrooms_frame, from_=1, to=100, width=21)
        apart_update_tl_bedrooms_var = IntVar()
        apart_update_tl_bedrooms_check = ttk.Checkbutton(apart_update_tl_bedrooms_frame, text="Retrieve old data", variable=apart_update_tl_bedrooms_var, command=get_existing_bedrooms)
        apart_update_tl_bedrooms_frame.pack(expand=TRUE, fill=BOTH)
        apart_update_tl_bedrooms_label.grid(row=0, column=0, padx=(10,15))
        apart_update_tl_bedrooms_spin.grid(row=0, column=1)
        apart_update_tl_bedrooms_check.grid(row=1, column=1, sticky=W, pady=5, padx=10)


        def get_existing_bathrooms():
            if apart_update_tl_bathroom_var.get() == 0:
                apart_update_tl_bathroom_spin.delete(0, END)
            else:
                apart_update_tl_bathroom_spin.delete(0, END)
                apart_update_tl_bathroom_spin.insert(0, update_tl_apart_l[6])
        apart_update_tl_bathroom_frame = Frame(apart_update_tl_rigth_frame)
        apart_update_tl_bathroom_label = ttk.Label(apart_update_tl_bathroom_frame, text="Bathroom:", font=("", 13, "bold"))
        apart_update_tl_bathroom_spin = ttk.Spinbox(apart_update_tl_bathroom_frame, from_=1, to=100, width=21)
        apart_update_tl_bathroom_var = IntVar()
        apart_update_tl_bathroom_check = ttk.Checkbutton(apart_update_tl_bathroom_frame, text="Retrieve old data", variable=apart_update_tl_bathroom_var, command=get_existing_bathrooms)
        apart_update_tl_bathroom_frame.pack(expand=TRUE, fill=BOTH)
        apart_update_tl_bathroom_label.grid(row=0, column=0, padx=(12,15))
        apart_update_tl_bathroom_spin.grid(row=0, column=1)
        apart_update_tl_bathroom_check.grid(row=1, column=1, sticky=W, pady=5, padx=10)


        def get_existing_kitchen():
            if apart_update_tl_kitchen_var.get() == 0:
                apart_update_tl_kitchen_combo.delete(0, END)
            else:
                apart_update_tl_kitchen_combo.delete(0, END)
                apart_update_tl_kitchen_combo.insert(0, update_tl_apart_l[7])
        apart_update_tl_kitchen_frame = Frame(apart_update_tl_rigth_frame)
        apart_update_tl_kitchen_label = ttk.Label(apart_update_tl_kitchen_frame, text="Kitchen:", font=("", 13, "bold"))
        apart_update_tl_kitchen_layouts = [
            "Galley Kitchen",
            "L-Shaped Kitchen",
            "U-Shaped Kitchen",
            "One-Wall Kitchen",
            "Peninsula Kitchen",
            "Island Kitchen"
        ]
        apart_update_tl_kitchen_var_s = StringVar()
        apart_update_tl_kitchen_combo = ttk.Combobox(apart_update_tl_kitchen_frame, textvariable=apart_update_tl_kitchen_var_s, values=apart_update_tl_kitchen_layouts, width=25)
        apart_update_tl_kitchen_var = IntVar()
        apart_update_tl_kitchen_check = ttk.Checkbutton(apart_update_tl_kitchen_frame, text="Retrieve old data", variable=apart_update_tl_kitchen_var, command=get_existing_kitchen)
        apart_update_tl_kitchen_frame.pack(expand=TRUE, fill=BOTH)
        apart_update_tl_kitchen_label.grid(row=0, column=0, padx=(10,33))
        apart_update_tl_kitchen_combo.grid(row=0, column=1)
        apart_update_tl_kitchen_check.grid(row=1, column=1, sticky=W, pady=5, padx=10)

        apart_optional_part_frame = Frame(apart_update_tl_bottom_frame_ente)


        def get_existing_livingroom():
            if apart_update_tl_livingroom_var.get() == 0:
                update_tl_apart_livingroom_var_s.set("")
            else:
                update_tl_apart_livingroom_var_s.set("")
                update_tl_apart_livingroom_var_s.set(update_tl_apart_l[9])
        update_tl_apart_livingroom_var_s = StringVar()
        apart_livingroom_label = ttk.Label(apart_optional_part_frame, text="Living Area:", font=("", 13, "bold"))
        included_living_btn = ttk.Radiobutton(apart_optional_part_frame, variable=update_tl_apart_livingroom_var_s, value='INCLUDED', text='INCLUDED')
        none_living_btn = ttk.Radiobutton(apart_optional_part_frame, variable=update_tl_apart_livingroom_var_s, value='NONE', text='NONE')
        apart_update_tl_livingroom_var = IntVar()
        apart_update_tl_livingroom_check = ttk.Checkbutton(apart_optional_part_frame, text="Retrieve old data", variable=apart_update_tl_livingroom_var, command=get_existing_livingroom)

        # Apartment Dining Area
        def get_existing_dining():
            if apart_update_tl_dining_var.get() == 0:
                update_tl_apart_dining_var_s.set("")
            else:
                update_tl_apart_dining_var_s.set("")
                update_tl_apart_dining_var_s.set(update_tl_apart_l[9])
        apart_dining_area_label = ttk.Label(apart_optional_part_frame, text="Dining Area:", font=("", 13, "bold"))
        update_tl_apart_dining_var_s = StringVar()
        included_dining_btn = ttk.Radiobutton(apart_optional_part_frame, variable=update_tl_apart_dining_var_s, value='INCLUDED', text='INCLUDED')
        none_dining_btn = ttk.Radiobutton(apart_optional_part_frame, variable=update_tl_apart_dining_var_s, value='NONE', text='NONE')
        apart_update_tl_dining_var = IntVar()
        apart_update_tl_dining_check = ttk.Checkbutton(apart_optional_part_frame, text="Retrieve old data", variable=apart_update_tl_dining_var, command=get_existing_dining)

        # Apartment Balcony
        def get_existing_balcony():
            if apart_update_tl_balcony_var.get() == 0:
                update_tl_apart_balcony_var_s.set("")
            else:
                update_tl_apart_balcony_var_s.set("")
                update_tl_apart_balcony_var_s.set(update_tl_apart_l[10])
        apart_balcony_label = ttk.Label(apart_optional_part_frame, text="Balcony:", font=("", 13, "bold"))
        update_tl_apart_balcony_var_s = StringVar()
        included_balcony_btn = ttk.Radiobutton(apart_optional_part_frame, variable=update_tl_apart_balcony_var_s, value='INCLUDED', text='INCLUDED')
        none_balcony_btn = ttk.Radiobutton(apart_optional_part_frame, variable=update_tl_apart_balcony_var_s, value='NONE', text='NONE')
        apart_update_tl_balcony_var = IntVar()
        apart_update_tl_balcony_check = ttk.Checkbutton(apart_optional_part_frame, text="Retrieve old data", variable=apart_update_tl_balcony_var, command=get_existing_balcony)

        # Apartment Storage
        def get_existing_storage():
            if apart_update_tl_storage_var.get() == 0:
                update_tl_apart_storage_var_s.set("")
            else:
                update_tl_apart_storage_var_s.set("")
                update_tl_apart_storage_var_s.set(update_tl_apart_l[11])
        apart_storage_label = ttk.Label(apart_optional_part_frame, text="Storage:", font=("", 13, "bold"))
        update_tl_apart_storage_var_s = StringVar()
        included_storange_btn = ttk.Radiobutton(apart_optional_part_frame, variable=update_tl_apart_storage_var_s, value='INCLUDED', text='INCLUDED')
        none_storange_btn = ttk.Radiobutton(apart_optional_part_frame, variable=update_tl_apart_storage_var_s, value='NONE', text='NONE')
        apart_update_tl_storage_var = IntVar()
        apart_update_tl_storage_check = ttk.Checkbutton(apart_optional_part_frame, text="Retrieve old data", variable=apart_update_tl_storage_var, command=get_existing_storage)

        # Apartment Laundry Room
        def get_existing_laundry():
            if apart_update_tl_laundry_var.get() == 0:
                update_tl_apart_laundry_var_s.set("")
            else:
                update_tl_apart_laundry_var_s.set("")
                update_tl_apart_laundry_var_s.set(update_tl_apart_l[12])
        apart_laundry_label = ttk.Label(apart_optional_part_frame, text="Laundry Room:", font=("", 13, "bold"))
        update_tl_apart_laundry_var_s = StringVar()
        included_laundry_btn = ttk.Radiobutton(apart_optional_part_frame, variable=update_tl_apart_laundry_var_s, value='INCLUDED', text='INCLUDED')
        none_laundry_btn = ttk.Radiobutton(apart_optional_part_frame, variable=update_tl_apart_laundry_var_s, value='NONE', text='NONE')
        apart_update_tl_laundry_var = IntVar()
        apart_update_tl_laundry_check = ttk.Checkbutton(apart_optional_part_frame, text="Retrieve old data", variable=apart_update_tl_laundry_var, command=get_existing_laundry)

        # Apartment Utilities
        apart_utilities_frame_main = ttk.LabelFrame(apart_update_tl_bottom_frame_ente, text="UTILITIES")
        apart_utilities_frame = Frame(apart_utilities_frame_main)
        # Apartment Utilities Water
        def update_apart_water():
            if update_tl_apart_water_var_s.get() == "" or update_tl_apart_water_var_s.get() == 'NONE':
                apart_update_tl_water_var.set(0)
                apart_update_tl_water_rate_frame.delete(0, END)
                apart_update_tl_water_rate_frame.insert(0, "Water")
                apart_update_tl_water_rate_frame['state'] = "disable"
            elif update_tl_apart_water_var_s.get() == 'INCLUDED':
                apart_update_tl_water_rate_frame['state'] = "normal"
                apart_update_tl_water_rate_frame.delete(0, END)
                if update_tl_apart_l[14] == "NONE":
                    apart_update_tl_water_rate_frame.insert(0, "Water")
                else:
                    apart_update_tl_water_rate_frame.insert(0, update_tl_apart_l[14])
        def get_existing_water():
            if apart_update_tl_water_var.get() == 0:
                update_tl_apart_water_var_s.set("")
                apart_update_tl_water_rate_frame.delete(0, END)
                apart_update_tl_water_rate_frame.insert(0, "Water")
                apart_update_tl_water_rate_frame['state'] = "disable"
            else:
                update_tl_apart_water_var_s.set("")
                update_tl_apart_water_var_s.set(update_tl_apart_l[13])
                apart_update_tl_water_rate_frame['state'] = "normal"
                apart_update_tl_water_rate_frame.delete(0, END)
                apart_update_tl_water_rate_frame.insert(0, update_tl_apart_l[14])

        apart_water_label = ttk.Label(apart_utilities_frame, text="Water:", font=("", 13, "bold"))
        update_tl_apart_water_var_s = StringVar()
        included_water_btn = ttk.Radiobutton(apart_utilities_frame, variable=update_tl_apart_water_var_s, value='INCLUDED', text='INCLUDED', command=update_apart_water)
        none_water_btn = ttk.Radiobutton(apart_utilities_frame, variable=update_tl_apart_water_var_s, value='NONE', text='NONE', command=update_apart_water)
        apart_update_tl_water_var = IntVar()
        apart_update_tl_water_check = ttk.Checkbutton(apart_utilities_frame, text="Retrieve old data", variable=apart_update_tl_water_var, command=get_existing_water)

        # Apartment Utilities Electricity      
        def update_apart_elec():
            if update_tl_apart_electricity_var_s.get() == "" or update_tl_apart_electricity_var_s.get() == 'NONE':
                apart_update_tl_electricity_var.set(0)
                apart_update_tl_elec_rate_frame.delete(0, END)
                apart_update_tl_elec_rate_frame.insert(0, "Electricity")
                apart_update_tl_elec_rate_frame['state'] = "disable"
            elif update_tl_apart_electricity_var_s.get() == 'INCLUDED':
                apart_update_tl_elec_rate_frame['state'] = "normal"
                apart_update_tl_elec_rate_frame.delete(0, END)
                if update_tl_apart_l[16] == "NONE":
                    apart_update_tl_elec_rate_frame.insert(0, "Electricity")
                else:
                    apart_update_tl_elec_rate_frame.insert(0, update_tl_apart_l[16])
        def get_existing_electricity():
            if apart_update_tl_electricity_var.get() == 0:
                update_tl_apart_electricity_var_s.set("")
                apart_update_tl_elec_rate_frame.delete(0, END)
                apart_update_tl_elec_rate_frame.insert(0, "Electricity")
                apart_update_tl_elec_rate_frame['state'] = "disable"
            else:
                update_tl_apart_electricity_var_s.set("")
                update_tl_apart_electricity_var_s.set(update_tl_apart_l[15])
                apart_update_tl_elec_rate_frame['state'] = "normal"
                apart_update_tl_elec_rate_frame.delete(0, END)
                apart_update_tl_elec_rate_frame.insert(0, update_tl_apart_l[16])
        apart_electricity_label = ttk.Label(apart_utilities_frame, text="Electricity:", font=("", 13, "bold"))
        update_tl_apart_electricity_var_s = StringVar()
        included_electricity_btn = ttk.Radiobutton(apart_utilities_frame, variable=update_tl_apart_electricity_var_s, value='INCLUDED', text='INCLUDED', command=update_apart_elec)
        none_electricity_btn = ttk.Radiobutton(apart_utilities_frame, variable=update_tl_apart_electricity_var_s, value='NONE', text='NONE', command=update_apart_elec)
        apart_update_tl_electricity_var = IntVar()
        apart_update_tl_electricity_check = ttk.Checkbutton(apart_utilities_frame, text="Retrieve old data", variable=apart_update_tl_electricity_var, command=get_existing_electricity)

        # Apartment Utilities Internet
        def update_apart_inter():
            if update_tl_apart_internet_var_s.get() == "" or update_tl_apart_internet_var_s.get() == 'NONE':
                apart_update_tl_internet_var.set(0)
                apart_update_tl_inter_rate_frame.delete(0, END)
                apart_update_tl_inter_rate_frame.insert(0, "Internet")
                apart_update_tl_inter_rate_frame['state'] = "disable"
            elif update_tl_apart_internet_var_s.get() == 'INCLUDED':
                apart_update_tl_inter_rate_frame['state'] = "normal"
                apart_update_tl_inter_rate_frame.delete(0, END)
                if update_tl_apart_l[18] == "NONE":
                    apart_update_tl_inter_rate_frame.insert(0, "Internet")
                else:
                    apart_update_tl_inter_rate_frame.insert(0, update_tl_apart_l[18])
        def get_existing_internet():
            if apart_update_tl_internet_var.get() == 0:
                update_tl_apart_internet_var_s.set("")
                apart_update_tl_inter_rate_frame.delete(0, END)
                apart_update_tl_inter_rate_frame.insert(0, "Internet")
                apart_update_tl_inter_rate_frame['state'] = "disable"
            else:
                update_tl_apart_internet_var_s.set("")
                update_tl_apart_internet_var_s.set(update_tl_apart_l[17])
                apart_update_tl_inter_rate_frame['state'] = "normal"
                apart_update_tl_inter_rate_frame.delete(0, END)
                apart_update_tl_inter_rate_frame.insert(0, update_tl_apart_l[18])
        apart_internet_label = ttk.Label(apart_utilities_frame, text="Internet:", font=("", 13, "bold"))
        update_tl_apart_internet_var_s = StringVar()
        included_internet_btn = ttk.Radiobutton(apart_utilities_frame, variable=update_tl_apart_internet_var_s, value='INCLUDED', text='INCLUDED', command=update_apart_inter)
        none_internet_btn = ttk.Radiobutton(apart_utilities_frame, variable=update_tl_apart_internet_var_s, value='NONE', text='NONE', command=update_apart_inter)
        apart_update_tl_internet_var = IntVar()
        apart_update_tl_internet_check = ttk.Checkbutton(apart_utilities_frame, text="Retrieve old data", variable=apart_update_tl_internet_var, command=get_existing_internet)


        apart_update_tl_utilities_rate_frame = ttk.LabelFrame(apart_utilities_frame_main, text="UTILITIES RATE")
        apart_update_tl_water_rate_frame = ttk.Entry(apart_update_tl_utilities_rate_frame, font=("", 13, "bold"), width=11)
        apart_update_tl_water_rate_frame.insert(0, "Water")
        apart_update_tl_water_rate_frame['state'] = "disable"
        def apart_update_tl_water_entry_i(event):
            if apart_update_tl_water_rate_frame.get().strip() == "" or \
                apart_update_tl_water_rate_frame.get().strip() == "Water":
                apart_update_tl_water_rate_frame.delete(0, END)
        def apart_update_tl_water_entry_o(event):
            if apart_update_tl_water_rate_frame.get().strip() == "":
                apart_update_tl_water_rate_frame.insert(0, "Water")
        apart_update_tl_water_rate_frame.bind("<FocusIn>", apart_update_tl_water_entry_i)
        apart_update_tl_water_rate_frame.bind("<FocusOut>", apart_update_tl_water_entry_o)

        apart_update_tl_elec_rate_frame  = ttk.Entry(apart_update_tl_utilities_rate_frame, font=("", 13, "bold"), width=11)
        apart_update_tl_elec_rate_frame.insert(0, "Electricity")
        apart_update_tl_elec_rate_frame['state'] = "disable"
        def apart_update_tl_Electricity_entry_i(event):
            if apart_update_tl_elec_rate_frame.get().strip() == "" or \
                apart_update_tl_elec_rate_frame.get().strip() == "Electricity":
                apart_update_tl_elec_rate_frame.delete(0, END)
        def apart_update_tl_Electricity_entry_o(event):
            if apart_update_tl_elec_rate_frame.get().strip() == "":
                apart_update_tl_elec_rate_frame.insert(0, "Electricity")
        apart_update_tl_elec_rate_frame.bind("<FocusIn>", apart_update_tl_Electricity_entry_i)
        apart_update_tl_elec_rate_frame.bind("<FocusOut>", apart_update_tl_Electricity_entry_o)

        apart_update_tl_inter_rate_frame = ttk.Entry(apart_update_tl_utilities_rate_frame, font=("", 13, "bold"), width=11)
        apart_update_tl_inter_rate_frame.insert(0, "Internet")
        apart_update_tl_inter_rate_frame['state'] = "disable"
        def apart_update_tl_inter_entry_i(event):
            if apart_update_tl_inter_rate_frame.get().strip() == "" or \
                apart_update_tl_inter_rate_frame.get().strip() == "Internet":
                apart_update_tl_inter_rate_frame.delete(0, END)
        def apart_update_tl_inter_entry_o(event):
            if apart_update_tl_inter_rate_frame.get().strip() == "":
                apart_update_tl_inter_rate_frame.insert(0, "Internet")
        apart_update_tl_inter_rate_frame.bind("<FocusIn>", apart_update_tl_inter_entry_i)
        apart_update_tl_inter_rate_frame.bind("<FocusOut>", apart_update_tl_inter_entry_o)


        apart_optional_part_frame.pack(fill=X, padx=5, pady=(0, 5))

        apart_livingroom_label.grid(row=0, column=0, sticky=W, padx=(5, 10), pady=(10, 5))
        included_living_btn.grid(row=0, column=1, padx=(30, 20))
        none_living_btn.grid(row=0, column=2, padx=(0, 20))
        apart_update_tl_livingroom_check.grid(row=1, column=0, padx=10)

        apart_dining_area_label.grid(row=2, column=0, sticky=W, padx=(5, 10), pady=5)
        included_dining_btn.grid(row=2, column=1, padx=(30, 20))
        none_dining_btn.grid(row=2, column=2, padx=(0, 20))
        apart_update_tl_dining_check.grid(row=3, column=0)

        apart_balcony_label.grid(row=4, column=0, sticky=W, padx=(5, 10), pady=5)
        included_balcony_btn.grid(row=4, column=1, padx=(30, 20))
        none_balcony_btn.grid(row=4, column=2, padx=(0, 20))
        apart_update_tl_balcony_check.grid(row=5, column=0)

        apart_storage_label.grid(row=6, column=0, sticky=W, padx=(5, 10), pady=5)
        included_storange_btn.grid(row=6, column=1, padx=(30, 20))
        none_storange_btn.grid(row=6, column=2, padx=(0, 20))
        apart_update_tl_storage_check.grid(row=7, column=0)

        apart_laundry_label.grid(row=8, column=0, sticky=W, padx=(5, 10), pady=5)
        included_laundry_btn.grid(row=8, column=1, padx=(30, 20))
        none_laundry_btn.grid(row=8, column=2, padx=(0, 20))
        apart_update_tl_laundry_check.grid(row=9, column=0)

        apart_utilities_frame_main.pack(fill=X, padx=5)
        apart_utilities_frame.pack(fill=X, padx=5)
        apart_update_tl_utilities_rate_frame.pack(fill=X, padx=5, pady=(0, 5))
        apart_water_label.grid(row=0, column=0, sticky=W, padx=(5, 10), pady=(10, 5))
        included_water_btn.grid(row=0, column=1, padx=(30, 20))
        none_water_btn.grid(row=0, column=2)
        apart_update_tl_water_check.grid(row=1, column=0, padx=10)

        apart_electricity_label.grid(row=2, column=0, sticky=W, padx=(5, 10), pady=5)
        included_electricity_btn.grid(row=2, column=1, padx=(30, 20))
        none_electricity_btn.grid(row=2, column=2)
        apart_update_tl_electricity_check.grid(row=3, column=0)

        apart_internet_label.grid(row=4, column=0, sticky=W, padx=(5, 10), pady=5)
        included_internet_btn.grid(row=4, column=1, padx=(30, 20))
        none_internet_btn.grid(row=4, column=2)
        apart_update_tl_internet_check.grid(row=5, column=0, pady=(0, 10))
        apart_update_tl_water_rate_frame.pack(side=LEFT, padx=5, pady=5)
        apart_update_tl_elec_rate_frame.pack(side=LEFT, pady=5, fill=X)
        apart_update_tl_inter_rate_frame.pack(side=LEFT, padx=5, pady=5)

        apart_updete_tl_add_btn = Button(apart_update_tl_bottom_frame, text="UPDATE", relief='flat', font=("", 13, "bold"), bg='#217346', command=apart_update_tl_update_func)
        apart_updete_tl_add_btn.pack(fill=X, padx=5, pady=5)

        apart_exit_tl_add_btn = ttk.Button(apart_update_tl_frame, text="APARTMENT MANAGEMENT", command=apart_update_tl_exit_func)
        apart_exit_tl_add_btn.pack(fill=X, ipady=5, pady=(5, 0))
        apart_have_found()
        
        apart_update_tl.mainloop()


    def apart_main_exit():
        manage_apart.destroy()
        main_interface()


    def apart_to_search():
        def exit_apart_search():
            apart_search_tree_tl.destroy()
            
        search_unit_number = apart_search_entry.get().strip()
        if search_unit_number != "" and search_unit_number != "Unit #":
            search_unit_data_found = False
            for search_every_cell_apart in range(2, apart_sheet.max_row+1):
                if  search_unit_number == apart_sheet['A'+str(search_every_cell_apart)].value:
                    search_number_row = search_every_cell_apart
                    search_unit_data_found = True 
                    break
            if search_unit_data_found:
                apart_search_tree_tl = tk.Toplevel()
                apart_search_tree_tl.resizable(0,0)
                apart_search_tree_tl.title("Apartment Details")

                apart_search_tree_main_frame = Frame(apart_search_tree_tl)
                apart_search_tree_main_frame.pack(expand=TRUE, fill=BOTH, padx=10, pady=10)
                cols = ["1", "2"]
                apart_search_tree = ttk.Treeview(apart_search_tree_main_frame, show='headings', columns=cols, height=18)
                apart_search_tree.column("1", anchor=W, width=150)
                apart_search_tree.heading("1", text="", anchor=W)
                apart_search_tree.column("2", anchor=W, width=150)
                apart_search_tree.heading("2", text="", anchor=W)
                list_of_data = [] 
                unit_heading = apart_sheet[1]
                unit_data = apart_sheet[search_number_row]
                ind = 0
                for every_data in unit_data:
                    heading = unit_heading[ind]
                    list_of_data.append([heading.value, every_data.value])
                    ind += 1
                for each_data in list_of_data:
                    apart_search_tree.insert("", END, values=each_data)
                apart_search_tree.pack(expand=TRUE, fill=Y)
                exit_btn = ttk.Button(apart_search_tree_main_frame, text="EXIT", command=exit_apart_search)
                exit_btn.pack(fill=X, pady=(5, 0))
            else:
                apart_search_entry.delete(0, END)
                apart_search_entry.insert(0, "Unit #")
                messagebox.showerror("ERROR", f"There is no existing data for \"{search_unit_number}\".")
        else:
            messagebox.showerror('ERROR', "Must provide a Unit Number.")


    def apart_to_delete(search_unit_number, verify):
        for every_apart in range(2, main_sheet.max_row+1):
            if search_unit_number == main_sheet['A'+str(every_apart)].value:
                unit_main_row = every_apart
                break
        for search_every_cell_apart in range(2, apart_sheet.max_row+1):
            if  search_unit_number == apart_sheet['A'+str(search_every_cell_apart)].value:
                search_number_row = search_every_cell_apart
                break
        apart_delete_data = messagebox.askquestion("Verifying", f"You're about to delete all information about UNIT {search_unit_number}? \
                                                \nNote: ACTION WILL DELETE ALL DATA ABOUT UNIT {search_unit_number} ACROSS ALL MANAGEMENTS PERMANENTLY \
                                                \n\nCLICK \"yes\" TO CONTINUE", 
                                                icon="question")
        if apart_delete_data == "yes":
            verify.destroy()
            main_sheet.delete_rows(unit_main_row)
            apart_sheet.delete_rows(search_number_row)
            excel_connection.save(excel_file)
            apartment()
        else:
            verify.destroy()
            apartment()


    def apart_add_new():
        unit_number = apart_unit_number_entry.get().strip()
        unit_status = apart_unit_status_var.get().strip()
        unit_size = apart_size_entry.get().strip()
        unit_rate = apart_rate_entry.get().strip()
        unit_floor = apart_floor_spin.get().strip()
        unit_bedroom = apart_nBedroom_spin.get().strip()
        unit_bathroom = apart_nBathroom_spin.get().strip()
        unit_kitchen = kitchen_layouts_var.get().strip()
        unit_livingroom = apart_livingroom_var.get().strip()
        unit_dining = apart_dining_var.get().strip()
        unit_balcony = apart_balcony_var.get().strip()
        unit_storage = apart_storage_var.get().strip()
        unit_laundry = apart_laundry_var.get().strip()
        unit_water = apart_water_var.get().strip()
        if unit_water == "NONE":
            water_rate = "NONE"
        else:
            water_rate = apart_water_rate_frame.get().strip()
        unit_electricity = apart_electricity_var.get().strip()
        if unit_electricity == "NONE":
            elect_rate = "NONE"
        else:
            elect_rate = apart_elec_rate_frame.get().strip()
        unit_internet = apart_internet_var.get().strip()
        if unit_internet == "NONE":
            inter_rate = "NONE"
        else:
            inter_rate = apart_inter_rate_frame.get().strip()

        new_apart_info_list = [
            unit_number, 
            unit_rate,
            unit_status, 
            unit_size,
            unit_floor,
            unit_bedroom,
            unit_bathroom,
            unit_kitchen,
            unit_livingroom,
            unit_dining,
            unit_balcony,
            unit_storage,
            unit_laundry,
            unit_water,
            water_rate,
            unit_electricity,
            elect_rate,
            unit_internet,
            inter_rate
            ]
        
        new_main_interface_list = [
            unit_number,
            unit_status,
            "NONE",
            unit_rate
        ]
        
        apart_unit_num_list = list(apart_sheet["A"])
        apart_unit_num_found = False
        for unit_number_excel in apart_unit_num_list[1:]:
            if unit_number == unit_number_excel.value:
                apart_unit_num_found = True
                break
        if apart_unit_num_found:
            messagebox.showerror("ERROR", f"UNIT \"{unit_number}\" ALREADY EXIST!")
        else:
            apart_entries_defult = ["Unit #", "Size (m^2)", "Rate", "Floors", "Bedrooms", "Bathrooms"]
            is_entry_full = False
            is_var_full = False
            is_entry_uti_full = False
            if unit_number != "" and unit_number != apart_entries_defult[0] and \
                unit_size != "" and unit_size != apart_entries_defult[1] and \
                unit_rate != "" and unit_rate != apart_entries_defult[2] and \
                unit_floor != "" and unit_floor != apart_entries_defult[3] and \
                unit_bedroom != "" and unit_bedroom != apart_entries_defult[4] and \
                unit_bathroom != "" and unit_bathroom != apart_entries_defult[5]:
                is_entry_full = True
            if unit_status != "" and unit_kitchen != "" and unit_livingroom != "" and \
                unit_dining != "" and unit_balcony != "" and unit_storage != "" and \
                unit_laundry != "" and unit_water != "" and unit_electricity != "" and \
                unit_internet != "":
                is_var_full = True
            if water_rate != "Water" and water_rate != "" and \
                elect_rate  != "Electricity" and elect_rate  != "" and \
                inter_rate != "Internet" and inter_rate != "":
                is_entry_uti_full = True
            # Saving new data
            if is_entry_full and is_var_full and is_entry_uti_full:
                apart_u_sure_msg = f"""
Double check the information:

Unit Number: {unit_number}
Rent Rate: {unit_rate}
Unit Status: {unit_status}
Unit Size: {unit_size} m^2
Number of floors: {unit_floor}
Number of bedrooms: {unit_bedroom}
Number of bathrooms: {unit_bathroom}
Kithcen type: {unit_kitchen}
Living area: {unit_livingroom}
Dining area: {unit_dining}
Balcony: {unit_balcony}
Storage room: {unit_storage}
Laundry room: {unit_laundry}
Water: {unit_water}, RATE: {water_rate}
Electricity: {unit_electricity}, RATE: {elect_rate}
Internet: {unit_internet}, RATE: {inter_rate}

click "Yes" to continue to add.
"""
                apart_add_ans = messagebox.askquestion("Validating!", apart_u_sure_msg, icon="question")
                if apart_add_ans == "yes":
                    main_sheet.append(new_main_interface_list)
                    apart_sheet.append(new_apart_info_list)
                    excel_connection.save(excel_file)
                    apart_refresh_tree(_apart_treeview)
                    apart_reset_entries()
                    messagebox.showinfo("Success", "Apartment Successfully Added.")
                    if unit_status == "Occupied":
                        messagebox.showinfo("Add tenant", f"Add the New Tenant to the New Apartment.")
                        manage_apart.destroy()
                        tenant()
            elif is_entry_full != True or is_var_full != True or is_entry_uti_full!= True:
                messagebox.showerror("ERROR", f"MUST FILL ALL THE FIELDS!")


    def apart_reset_entries():
        apart_search_entry.delete(0, END)
        apart_search_entry.insert(0, "Unit #")
        apart_unit_number_entry.delete(0, END)
        apart_unit_number_entry.insert(0, "Unit #")
        apart_size_entry.delete(0, END)
        apart_size_entry.insert(0, "Size (m^2)")
        apart_rate_entry.delete(0, END)
        apart_rate_entry.insert(0, "Rate")
        apart_floor_spin.delete(0, END)
        apart_floor_spin.insert(0, "Floors")
        apart_nBedroom_spin.delete(0, END)
        apart_nBedroom_spin.insert(0, "Bedrooms")
        apart_nBathroom_spin.delete(0, END)
        apart_nBathroom_spin.insert(0, "Bathrooms")
        apart_livingroom_var.set("")
        apart_dining_var.set("")
        apart_balcony_var.set("")
        apart_storage_var.set("")
        apart_laundry_var.set("")
        apart_water_var.set("")
        apart_electricity_var.set("")
        apart_internet_var.set("")
        apart_water_rate_frame.delete(0, END)
        apart_water_rate_frame.insert(0, "Water")
        apart_water_rate_frame['state'] = "disable"
        apart_elec_rate_frame.delete(0, END)
        apart_elec_rate_frame.insert(0, "Electricity")
        apart_elec_rate_frame['state'] = "disable"
        apart_inter_rate_frame.delete(0, END)
        apart_inter_rate_frame.insert(0, "Internet")
        apart_inter_rate_frame['state'] = "disable"


    def manage_tenant_func():
        manage_apart.destroy()
        tenant()


    def manage_payment_func():
        manage_apart.destroy()
        payment()


    def apart_refresh_tree(apart_tree_refresh):
        apart_updated_data = list(apart_sheet.values)

        apart_tree_refresh.delete(*apart_tree_refresh.get_children())
        for every_row_apart in apart_updated_data[1:]:
            apart_tree_refresh.insert("", END, values=every_row_apart)


    apart_main_frame = Frame(manage_apart)
    apart_top_frame = Frame(apart_main_frame)
    apart_left_frame = Frame(apart_main_frame)
    apart_right_frame = Frame(apart_main_frame)
    apart_title = Label(apart_top_frame, font=("Arial Black", 70), text="Aparment Management",
                    fg="#217346")

    apart_search_frame = ttk.LabelFrame(apart_left_frame, text="SEARCH • UPDATE • DELETE")
    apart_search_entry = ttk.Entry(apart_search_frame, font=("", 13, "bold"), width=20)
    apart_search_entry.insert(0, "Unit #")
    def apart_search_entry_i(event):
        if apart_search_entry.get().strip() == "" or \
            apart_search_entry.get().strip() == "Unit #":
            apart_search_entry.delete(0, END)
    def apart_search_entry_o(event):
        if apart_search_entry.get().strip() == "":
            apart_search_entry.insert(0, "Unit #")
    apart_search_entry.bind("<FocusIn>", apart_search_entry_i)
    apart_search_entry.bind("<FocusOut>", apart_search_entry_o)

    apart_search_btn = Button(apart_search_frame, text="SEARCH", relief='flat', font=("", 13, "bold"), bg='#217346', command=apart_to_search)
    apart_update_btn = Button(apart_search_frame, text="UPDATE", relief='flat', font=("", 13, "bold"), bg='#217346', 
                              command=lambda:verify(management=manage_apart, 
                                                    action="updating_apartment", unit_to_update=apart_search_entry.get().strip(), 
                                                    function_for_action=apart_to_update))
    apart_delete_btn = ttk.Button(apart_search_frame, text="DELETE", command=lambda:verify(action="deleting_apartment", unit_to_delete=apart_search_entry.get(), management=manage_apart, function_for_action=apart_to_delete))


    apart_label_add_frame = Frame(apart_left_frame)
    apart_add_record = ttk.LabelFrame(apart_label_add_frame, text="ADD NEW UNIT")
    apart_big_add_frame = Frame(apart_add_record)
    apart_right_add_frame = Frame(apart_big_add_frame)

    # Status of Unit
    apart_unit_status = [
        "Under Construction",
        "Vacant",
        "Occupied"
    ]
    apart_unit_status_var = StringVar()
    apart_unit_status_combo = ttk.Combobox(apart_right_add_frame, textvariable=apart_unit_status_var, values=apart_unit_status, font=("", 13, "bold"), state="readonly")
    apart_unit_status_combo.current(1)
    # Unit Number and Aparment Size Frame
    # Apartment Unit Number
    apart_unit_number_entry = ttk.Entry(apart_right_add_frame, width=15, font=("", 13, "bold"))
    apart_unit_number_entry.insert(0, "Unit #")
    def apart_uN_entry_i(event):
        if apart_unit_number_entry.get().strip() == "" or \
            apart_unit_number_entry.get().strip() == "Unit #":
            apart_unit_number_entry.delete(0, END)
    def apart_uN_entry_o(event):
        if apart_unit_number_entry.get().strip() == "":
            apart_unit_number_entry.insert(0, "Unit #")
    apart_unit_number_entry.bind("<FocusIn>", apart_uN_entry_i)
    apart_unit_number_entry.bind("<FocusOut>", apart_uN_entry_o)
    # Apartment Size
    apart_size_entry = ttk.Entry(apart_right_add_frame, font=("", 13, "bold"))
    apart_size_entry.insert(0, "Size (m^2)")
    def apart_sz_entry_i(event):
        if apart_size_entry.get().strip() == "" or \
            apart_size_entry.get().strip() == "Size (m^2)":
            apart_size_entry.delete(0, END)
    def apart_sz_entry_o(event):
        if apart_size_entry.get().strip() == "":
            apart_size_entry.insert(0, "Size (m^2)")
    apart_size_entry.bind("<FocusIn>", apart_sz_entry_i)
    apart_size_entry.bind("<FocusOut>", apart_sz_entry_o)
    # Apartment Rate
    apart_rate_entry = ttk.Entry(apart_right_add_frame, font=("", 13, "bold"))
    apart_rate_entry.insert(0, "Rate")
    def apart_rate_sz_entry_i(event):
        if apart_rate_entry.get().strip() == "" or \
            apart_rate_entry.get().strip() == "Rate":
            apart_rate_entry.delete(0, END)
    def apart_rate_sz_entry_o(event):
        if apart_rate_entry.get().strip() == "":
            apart_rate_entry.insert(0, "Rate")
    apart_rate_entry.bind("<FocusIn>", apart_rate_sz_entry_i)
    apart_rate_entry.bind("<FocusOut>", apart_rate_sz_entry_o)


    # Spinboxes
    # Apartment Floors
    apart_floor_spin = ttk.Spinbox(apart_right_add_frame, from_=1, to=100, width=10, font=("", 13, "bold"))
    apart_floor_spin.insert(0, "Floors")
    def apart_fl_entry_i(event):
        if apart_floor_spin.get().strip() == "" or \
            apart_floor_spin.get().strip() == "Floors":
            apart_floor_spin.delete(0, END)
    def apart_fl_entry_o(event):
        if apart_floor_spin.get().strip() == "":
            apart_floor_spin.insert(0, "Floors")
    apart_floor_spin.bind("<FocusIn>", apart_fl_entry_i)
    apart_floor_spin.bind("<FocusOut>", apart_fl_entry_o)
    # Apartment Bedroom
    apart_nBedroom_spin = ttk.Spinbox(apart_right_add_frame, from_=1, to=100, width=10, font=("", 13, "bold"))
    apart_nBedroom_spin.insert(0, "Bedrooms")
    def apart_nb_i(event):
        if apart_nBedroom_spin.get().strip() == "" or \
            apart_nBedroom_spin.get().strip() == "Bedrooms":
            apart_nBedroom_spin.delete(0, END)
    def apart_nb_o(event):
        if apart_nBedroom_spin.get().strip() == "":
            apart_nBedroom_spin.insert(0, "Bedrooms")
    apart_nBedroom_spin.bind("<FocusIn>", apart_nb_i)
    apart_nBedroom_spin.bind("<FocusOut>", apart_nb_o)
    # Apartment Bathroom
    apart_nBathroom_spin = ttk.Spinbox(apart_right_add_frame, from_=1, to=100, width=10, font=("", 13, "bold"))
    apart_nBathroom_spin.insert(0, "Bathrooms")
    def apart_nb_entry_i(event):
        if apart_nBathroom_spin.get().strip() == "" or \
            apart_nBathroom_spin.get().strip() == "Bathrooms":
            apart_nBathroom_spin.delete(0, END)
    def apart_nb_entry_o(event):
        if apart_nBathroom_spin.get().strip() == "":
            apart_nBathroom_spin.insert(0, "Bathrooms")
    apart_nBathroom_spin.bind("<FocusIn>", apart_nb_entry_i)
    apart_nBathroom_spin.bind("<FocusOut>", apart_nb_entry_o)

    # Apartment Kitchen
    kitchen_layouts = [
        "Galley Kitchen",
        "L-Shaped Kitchen",
        "U-Shaped Kitchen",
        "One-Wall Kitchen",
        "Peninsula Kitchen",
        "Island Kitchen"
    ]
    kitchen_layouts_var = StringVar()
    apart_kitchen_type_combo = ttk.Combobox(apart_right_add_frame, textvariable=kitchen_layouts_var, values=kitchen_layouts, font=("", 13, "bold"), state="readonly")
    apart_kitchen_type_combo.current(1)
    # Apartment Living Room
    apart_left_add_frame = Frame(apart_big_add_frame)
    apart_optional_part_frame = Frame(apart_left_add_frame)
    apart_livingroom_var = StringVar()
    apart_livingroom_label = ttk.Label(apart_optional_part_frame, text="Living Area:", font=("", 13, "bold"))
    included_living_btn = ttk.Radiobutton(apart_optional_part_frame, variable=apart_livingroom_var, value='INCLUDED', text='INCLUDED')
    none_living_btn = ttk.Radiobutton(apart_optional_part_frame, variable=apart_livingroom_var, value='NONE', text='NONE')
    # Apartment Dining Area=
    apart_dining_area_label = ttk.Label(apart_optional_part_frame, text="Dining Area:", font=("", 13, "bold"))
    apart_dining_var = StringVar()
    included_dining_btn = ttk.Radiobutton(apart_optional_part_frame, variable=apart_dining_var, value='INCLUDED', text='INCLUDED')
    none_dining_btn = ttk.Radiobutton(apart_optional_part_frame, variable=apart_dining_var, value='NONE', text='NONE')
    # Apartment Balcony
    apart_balcony_label = ttk.Label(apart_optional_part_frame, text="Balcony:", font=("", 13, "bold"))
    apart_balcony_var = StringVar()
    included_balcony_btn = ttk.Radiobutton(apart_optional_part_frame, variable=apart_balcony_var, value='INCLUDED', text='INCLUDED')
    none_balcony_btn = ttk.Radiobutton(apart_optional_part_frame, variable=apart_balcony_var, value='NONE', text='NONE')
    # Apartment Storage
    apart_storage_label = ttk.Label(apart_optional_part_frame, text="Storage:", font=("", 13, "bold"))
    apart_storage_var = StringVar()
    included_storange_btn = ttk.Radiobutton(apart_optional_part_frame, variable=apart_storage_var, value='INCLUDED', text='INCLUDED')
    none_storange_btn = ttk.Radiobutton(apart_optional_part_frame, variable=apart_storage_var, value='NONE', text='NONE')
    # Apartment Landry Room
    apart_laundry_label = ttk.Label(apart_optional_part_frame, text="Laundry Room:", font=("", 13, "bold"))
    apart_laundry_var = StringVar()
    included_laundry_btn = ttk.Radiobutton(apart_optional_part_frame, variable=apart_laundry_var, value='INCLUDED', text='INCLUDED')
    none_laundry_btn = ttk.Radiobutton(apart_optional_part_frame, variable=apart_laundry_var, value='NONE', text='NONE')

    # Apartment Utilities
    apart_utilities_frame = ttk.LabelFrame(apart_left_add_frame, text="UTILITIES")
    apart_utilities_frame_top = Frame(apart_utilities_frame)
    # Apartment Utilities Water
    def apart_water():
        if apart_water_var.get() == "" or apart_water_var.get() == 'NONE':
            apart_water_rate_frame.delete(0, END)
            apart_water_rate_frame.insert(0, "Water")
            apart_water_rate_frame['state'] = "disable"
        elif apart_water_var.get() == 'INCLUDED':
            apart_water_rate_frame['state'] = "normal"
    apart_water_label = ttk.Label(apart_utilities_frame_top, text="Water:", font=("", 13, "bold"))
    apart_water_var = StringVar()
    included_water_btn = ttk.Radiobutton(apart_utilities_frame_top, variable=apart_water_var, value='INCLUDED', text='INCLUDED', command=apart_water)
    none_water_btn = ttk.Radiobutton(apart_utilities_frame_top, variable=apart_water_var, value='NONE', text='NONE', command=apart_water)
    # Apartment Utilities Electricity
    def apart_elect():
        if apart_electricity_var.get() == "" or apart_electricity_var.get() == 'NONE':
            apart_elec_rate_frame.delete(0, END)
            apart_elec_rate_frame.insert(0, "Electricity")
            apart_elec_rate_frame['state'] = "disable"
        elif apart_electricity_var.get() == 'INCLUDED':
            apart_elec_rate_frame['state'] = "normal"
    apart_electricity_label = ttk.Label(apart_utilities_frame_top, text="Electricity:", font=("", 13, "bold"))
    apart_electricity_var = StringVar()
    included_electricity_btn = ttk.Radiobutton(apart_utilities_frame_top, variable=apart_electricity_var, value='INCLUDED', text='INCLUDED', command=apart_elect)
    none_electricity_btn = ttk.Radiobutton(apart_utilities_frame_top, variable=apart_electricity_var, value='NONE', text='NONE', command=apart_elect)
    # Apartment Utilities Internet
    def apart_inter():
        if apart_internet_var.get() == "" or apart_internet_var.get() == 'NONE':
            apart_inter_rate_frame.delete(0, END)
            apart_inter_rate_frame.insert(0, "Internet")
            apart_inter_rate_frame['state'] = "disable"
        elif apart_internet_var.get() == 'INCLUDED':
            apart_inter_rate_frame['state'] = "normal"
    apart_internet_label = ttk.Label(apart_utilities_frame_top, text="Internet:", font=("", 13, "bold"))
    apart_internet_var = StringVar()
    included_internet_btn = ttk.Radiobutton(apart_utilities_frame_top, variable=apart_internet_var, value='INCLUDED', text='INCLUDED', command=apart_inter)
    none_internet_btn = ttk.Radiobutton(apart_utilities_frame_top, variable=apart_internet_var, value='NONE', text='NONE', command=apart_inter)

    # Utilities Rate
    apart_utilities_rate_label_frame = ttk.LabelFrame(apart_utilities_frame, text="UTILITIES RATE")
    apart_water_rate_frame = ttk.Entry(apart_utilities_rate_label_frame, font=("", 13, "bold"), width=8)
    apart_water_rate_frame.insert(0, "Water")
    apart_water_rate_frame['state'] = "disable"
    def apart_water_entry_i(event):
        if apart_water_rate_frame.get().strip() == "" or \
            apart_water_rate_frame.get().strip() == "Water":
            apart_water_rate_frame.delete(0, END)
    def apart_water_entry_o(event):
        if apart_water_rate_frame.get().strip() == "":
            apart_water_rate_frame.insert(0, "Water")
    apart_water_rate_frame.bind("<FocusIn>", apart_water_entry_i)
    apart_water_rate_frame.bind("<FocusOut>", apart_water_entry_o)

    apart_elec_rate_frame  = ttk.Entry(apart_utilities_rate_label_frame, font=("", 13, "bold"), width=8)
    apart_elec_rate_frame.insert(0, "Electricity")
    apart_elec_rate_frame['state'] = "disable"
    def apart_Electricity_entry_i(event):
        if apart_elec_rate_frame.get().strip() == "" or \
            apart_elec_rate_frame.get().strip() == "Electricity":
            apart_elec_rate_frame.delete(0, END)
    def apart_Electricity_entry_o(event):
        if apart_elec_rate_frame.get().strip() == "":
            apart_elec_rate_frame.insert(0, "Electricity")
    apart_elec_rate_frame.bind("<FocusIn>", apart_Electricity_entry_i)
    apart_elec_rate_frame.bind("<FocusOut>", apart_Electricity_entry_o)

    apart_inter_rate_frame = ttk.Entry(apart_utilities_rate_label_frame, font=("", 13, "bold"), width=8)
    apart_inter_rate_frame.insert(0, "Internet")
    apart_inter_rate_frame['state'] = "disable"
    def apart_inter_entry_i(event):
        if apart_inter_rate_frame.get().strip() == "" or \
            apart_inter_rate_frame.get().strip() == "Internet":
            apart_inter_rate_frame.delete(0, END)
    def apart_inter_entry_o(event):
        if apart_inter_rate_frame.get().strip() == "":
            apart_inter_rate_frame.insert(0, "Internet")
    apart_inter_rate_frame.bind("<FocusIn>", apart_inter_entry_i)
    apart_inter_rate_frame.bind("<FocusOut>", apart_inter_entry_o)


    apart_add_btn = Button(apart_add_record, text="ADD APARTMENT", relief='flat', font=("", 13, "bold"), bg='#217346', command=apart_add_new)

    apart_other_management_frame = Frame(apart_left_frame)
    apart_tennat_man_btn = ttk.Button(apart_other_management_frame, text="MANAGE TENANTS", command=manage_tenant_func)
    apart_paym_btn = ttk.Button(apart_other_management_frame, text="MANAGE PAYMENTS", command=manage_payment_func)


    apart_tree_bottom_frame = Frame(apart_right_frame) 

    excel_items_tree = list(apart_sheet.values)
    _apart_treeview = ttk.Treeview(apart_tree_bottom_frame, show='headings', columns=excel_items_tree[0])
    apart_scroll_bar_y = ttk.Scrollbar(apart_tree_bottom_frame, orient='vertical', command=_apart_treeview.yview)
    apart_scroll_bar_x = ttk.Scrollbar(apart_tree_bottom_frame, orient='horizontal', command=_apart_treeview.xview)
    _apart_treeview.configure(yscrollcommand=apart_scroll_bar_y.set,xscrollcommand=apart_scroll_bar_x.set)
    for heading in excel_items_tree[0]:
        if heading == "UNIT":
            _apart_treeview.column(heading, anchor=CENTER, width=40)
            _apart_treeview.heading(heading, text=heading)
        elif heading == "FLOORS":
            _apart_treeview.column(heading, anchor=CENTER, width=70)
            _apart_treeview.heading(heading, text=heading)
        elif heading == "STATUS":
            _apart_treeview.column(heading, anchor=CENTER, width=70)
            _apart_treeview.heading(heading, text=heading)
        elif heading == "BATHROOM":
            _apart_treeview.column(heading, anchor=CENTER, width=80)
            _apart_treeview.heading(heading, text=heading)
        elif heading == "BEDROOMS":
            _apart_treeview.column(heading, anchor=CENTER, width=90)
            _apart_treeview.heading(heading, text=heading)
        elif heading == "STATUS":
            _apart_treeview.column(heading, anchor=CENTER)
            _apart_treeview.heading(heading, text=heading)
        else:
            _apart_treeview.column(heading, anchor=CENTER, width=150)
            _apart_treeview.heading(heading, text=heading)
    for value in excel_items_tree[1:]:
        _apart_treeview.insert('', END, values=value)
    exit_apart_management = ttk.Button(apart_right_frame, text="MAIN PAGE", command=apart_main_exit)

    apart_main_frame.pack(padx=10, pady=(0,10) , expand=TRUE, fill=BOTH)
    apart_top_frame.pack(fill=X) 
    apart_title.pack()
    apart_left_frame.pack(expand=TRUE, side=LEFT)  
    apart_right_frame.pack(expand=TRUE,padx=(5, 0), side=RIGHT, fill=BOTH, pady=(7, 0))  

    apart_search_frame.pack(expand=TRUE, fill=BOTH, pady=(0, 10))
    apart_search_entry.pack(fill=X, pady=5, padx=5)
    apart_search_btn.pack(expand=TRUE, fill=X, side=LEFT, pady=5, padx=(5, 0))
    apart_update_btn.pack(expand=TRUE, fill=X, side=LEFT, pady=5, padx=(5, 0))
    apart_delete_btn.pack(expand=TRUE, fill=X, side=LEFT, pady=5, padx=5)

    apart_label_add_frame.pack(expand=TRUE, fill=BOTH)  
    apart_add_record.pack(expand=TRUE, fill=BOTH)
    apart_big_add_frame.pack(expand=TRUE, fill=BOTH)
    apart_right_add_frame.pack(side=LEFT, expand=TRUE, fill=BOTH)
    apart_left_add_frame.pack(side=LEFT, expand=TRUE, fill=BOTH)
    apart_unit_number_entry.pack(fill=X, padx=5, pady=(10, 8))
    apart_unit_status_combo.pack(fill=X, padx=5, pady=(0, 5))
    apart_size_entry.pack(expand=TRUE, fill=X, padx=5, pady=(0, 5))
    apart_rate_entry.pack(expand=TRUE, fill=X, padx=5, pady=(0, 7))

    apart_floor_spin.pack(fill=X, padx=5, pady=(0, 10))
    apart_nBedroom_spin.pack(fill=X, padx=5, pady=(0, 10))
    apart_nBathroom_spin.pack(fill=X, padx=5, pady=(0, 10))

    apart_kitchen_type_combo.pack(fill=X, padx=5)

    apart_optional_part_frame.pack(fill=X, padx=5, pady=(0, 5))
    apart_livingroom_label.grid(row=0, column=0, sticky=W, padx=(5, 10), pady=(10, 5))
    included_living_btn.grid(row=0, column=1, padx=(10, 15))
    none_living_btn.grid(row=0, column=2, padx=(0, 20))

    apart_dining_area_label.grid(row=1, column=0, sticky=W, padx=(5, 10), pady=5)
    included_dining_btn.grid(row=1, column=1, padx=(10, 15))
    none_dining_btn.grid(row=1, column=2, padx=(0, 20))

    apart_balcony_label.grid(row=2, column=0, sticky=W, padx=(5, 10), pady=5)
    included_balcony_btn.grid(row=2, column=1, padx=(10, 15))
    none_balcony_btn.grid(row=2, column=2, padx=(0, 20))

    apart_storage_label.grid(row=3, column=0, sticky=W, padx=(5, 10), pady=5)
    included_storange_btn.grid(row=3, column=1, padx=(10, 15))
    none_storange_btn.grid(row=3, column=2, padx=(0, 20))

    apart_laundry_label.grid(row=4, column=0, sticky=W, padx=(5, 10), pady=5)
    included_laundry_btn.grid(row=4, column=1, padx=(10, 15))
    none_laundry_btn.grid(row=4, column=2, padx=(0, 20))

    apart_utilities_frame.pack(fill=X, padx=5)
    apart_utilities_frame_top.pack(anchor=W)
    apart_water_label.grid(row=0, column=0, sticky=W, padx=(5, 10), pady=(10, 5))
    included_water_btn.grid(row=0, column=1, padx=(47, 15), sticky=W)
    none_water_btn.grid(row=0, column=2, sticky=W)

    apart_electricity_label.grid(row=1, column=0, sticky=W, padx=(5, 10), pady=5)
    included_electricity_btn.grid(row=1, column=1, padx=(47, 15), sticky=W)
    none_electricity_btn.grid(row=1, column=2, sticky=W)

    apart_internet_label.grid(row=2, column=0, sticky=W, padx=(5, 10), pady=5)
    included_internet_btn.grid(row=2, column=1, padx=(47, 15), pady=(0, 5), sticky=W)
    none_internet_btn.grid(row=2, column=2, sticky=W)

    apart_utilities_rate_label_frame.pack(anchor=W, fill=X, padx=5, pady=5)
    apart_water_rate_frame.pack(expand=TRUE, fill=X, side=LEFT, padx=5, pady=5)
    apart_elec_rate_frame.pack(expand=TRUE, fill=X, side=LEFT, pady=5)
    apart_inter_rate_frame.pack(expand=TRUE, fill=X, side=LEFT, padx=5, pady=5)

    apart_add_btn.pack(fill=X, padx=5, pady=5)

    apart_tree_bottom_frame.pack(expand=TRUE, fill=BOTH, pady=(2,0))
    apart_scroll_bar_y.pack(side=RIGHT, fill=Y)
    _apart_treeview.pack(expand=TRUE, fill=Y)
    apart_scroll_bar_x.pack(side=BOTTOM, fill=X)

    apart_other_management_frame.pack(fill=X, pady=(5,0))
    apart_tennat_man_btn.pack(fill=BOTH, expand=TRUE, side=LEFT, ipady=3)
    apart_paym_btn.pack(fill=BOTH, expand=TRUE, side=LEFT, padx=(5, 0), ipady=3)
    exit_apart_management.pack(fill=X, pady=(5, 0), ipady=3)

    manage_apart.mainloop()


def tenant():
    manage_tenants = tk.Tk()
    manage_tenants.title("Tenants Management")
    manage_tenants.resizable(0, 0)
    # Theme
    style = ttk.Style(manage_tenants)
    manage_tenants.tk.call("source", 'theme_forest-dark.tcl')
    style.theme_use("forest-dark")


    def is_unit_exist(unit_num):
        found_apartment = False
        unit_data_l = list(apart_sheet["A"])
        for every_unit_num in unit_data_l[1:]:
            if unit_num == every_unit_num.value:
                found_apartment = True
                break
        return found_apartment


    def tenant_to_search():
        def exit_tenant_search():
            tenant_search_tree_tl.destroy()
            
        search_unit_number = tenant_search_entry.get().strip()
        if search_unit_number != "" and search_unit_number != "Unit #":
            search_unit_data_found = False
            for search_every_cell in range(2, tenant_main_sheet.max_row+1):
                if  search_unit_number == tenant_main_sheet['A'+str(search_every_cell)].value:
                    search_number_row = search_every_cell
                    search_unit_data_found = True 
                    break
            if search_unit_data_found:
                tenant_search_tree_tl = tk.Toplevel()
                tenant_search_tree_tl.resizable(0,0)
                tenant_search_tree_tl.title("Tenants Details")

                tenant_search_tree_main_frame = Frame(tenant_search_tree_tl)
                tenant_search_tree_main_frame.pack(expand=TRUE, fill=BOTH, padx=10, pady=10)
                cols = ["1", "2"]
                tenant_search_tree = ttk.Treeview(tenant_search_tree_main_frame, show='headings', columns=cols, height=7)
                tenant_search_tree.column("1", anchor=W, width=150)
                tenant_search_tree.heading("1", text="")
                tenant_search_tree.column("2", anchor=W, width=150)
                tenant_search_tree.heading("2", text="DATA", anchor=W)
                list_of_data = [] 
                unit_heading = tenant_main_sheet[1]
                unit_data = tenant_main_sheet[search_number_row]
                ind = 0
                for every_data in unit_data:
                    heading = unit_heading[ind]
                    list_of_data.append([heading.value, every_data.value])
                    ind += 1
                for each_data in list_of_data:
                    tenant_search_tree.insert("", END, values=each_data)
                tenant_search_tree.pack(expand=TRUE, fill=Y)
                exit_btn = ttk.Button(tenant_search_tree_main_frame, text="EXIT", command=exit_tenant_search)
                exit_btn.pack(fill=X, pady=(5, 0))
            else:
                tenant_search_entry.delete(0, END)
                tenant_search_entry.insert(0, "Unit #")
                messagebox.showerror("ERROR", f"There is no existing data for \"{search_unit_number}\".")
        else:
            messagebox.showerror('ERROR', "Must provide a Unit Number.")


    def tenant_to_update(to_update_unit_num, verify):
        verify.destroy()
        tenant_row = 1
        current_unit_tenant_data = list(tenant_main_sheet["A"])
        for every_tenant_data in current_unit_tenant_data[1:]:
            tenant_row += 1
            if to_update_unit_num == every_tenant_data.value:
                current_unit = every_tenant_data.value
                break
        update_tl_tenant_l = []
        update_tl_tenant_sheet_l = list(tenant_main_sheet[tenant_row])
        for update_tl_tenant_every_item in update_tl_tenant_sheet_l:
            update_tl_tenant_l.append(update_tl_tenant_every_item.value)

        manage_updete_tl_tenants = tk.Tk()
        manage_updete_tl_tenants.title("Tenants Management")
        # Theme
        style = ttk.Style(manage_updete_tl_tenants)
        manage_updete_tl_tenants.tk.call("source", 'theme_forest-dark.tcl')
        style.theme_use("forest-dark")
        def update_tl_exit():
            manage_updete_tl_tenants.destroy()
            tenant()
        

        def to_update_tenant():
            new_tenant_unit = tenant_updete_tl_new_unit_number_enrty.get().strip()
            unit_for_tenant_found = False
            unit_row = 1 
            new_unit_tenant_data = list(apart_sheet["A"])
            for every_unit_data in new_unit_tenant_data[1:]:
                unit_row += 1
                if new_tenant_unit == every_unit_data.value:
                    unit_for_tenant_found = True
                    break
                
            if unit_for_tenant_found:
                is_new_unit_vacant = True
                if new_tenant_unit != current_unit:
                    if  apart_sheet[f"C{unit_row}"].value == "Occupied":
                        is_new_unit_vacant = False
                if is_new_unit_vacant:
                    u_sure = messagebox.askquestion("Are you sure?", "Your about to update tenant's data.\n\nNote: Action will make change across all management\n\nClick \"Yes\" to continue.")
                    if u_sure == "yes":
                        d_r = 0
                        tenant_hd = False
                        for e_due in range(2, payment_main_sheet.max_row+1):
                            if new_tenant_unit == payment_main_sheet["A"+str(e_due)].value:
                                d_r = e_due
                                tenant_hd = True
                                break
                        if tenant_updete_tl_new_unit_number_enrty.get().strip() == "":
                            updated_tenant_unit = tenant_main_sheet[f"A{tenant_row}"].value
                        else:
                            updated_tenant_unit = tenant_updete_tl_new_unit_number_enrty.get().strip()
                        if tenant_updete_tl_new_name_entry.get().strip() == "":
                            updated_tenant_name = tenant_main_sheet[f"B{tenant_row}"].value
                        else:
                            updated_tenant_name = tenant_updete_tl_new_name_entry.get().strip()
                        if tenant_updete_tl_new_age_entry.get().strip() == "":
                            updated_tenant_age = tenant_main_sheet[f"C{tenant_row}"].value
                        else:
                            updated_tenant_age = tenant_updete_tl_new_age_entry.get().strip()
                        if tenant_updete_tl_new_occupation_entry.get().strip() == "":
                            updated_tenant_occu = tenant_main_sheet[f"D{tenant_row}"].value
                        else:
                            updated_tenant_occu = tenant_updete_tl_new_occupation_entry.get().strip()
                        if tenant_updete_tl_new_gender_var.get().strip() == "":
                            updated_tenant_gender = tenant_main_sheet[f"E{tenant_row}"].value
                        else:
                            updated_tenant_gender = tenant_updete_tl_new_gender_var.get().strip()
                        if tenant_updete_tl_phone_number.get().strip() == "":
                            updated_tenant_phone = tenant_main_sheet[f"F{tenant_row}"].value
                        else:
                            updated_tenant_phone = tenant_updete_tl_phone_number.get().strip()
                        if tenant_updete_tl_email_entry.get().strip() == "":
                            updated_tenant_email = tenant_main_sheet[f"G{tenant_row}"].value
                        else:
                            updated_tenant_email = tenant_updete_tl_email_entry.get().strip()
                        for every_apart in range(2, main_sheet.max_row+1):
                            if new_tenant_unit == main_sheet['A'+str(every_apart)].value:
                                unit_main_row = every_apart
                                break
                        main_sheet["C"+str(unit_main_row)] = updated_tenant_name
                        tenant_main_sheet[f"A{tenant_row}"] = updated_tenant_unit
                        tenant_main_sheet[f"B{tenant_row}"] = updated_tenant_name
                        tenant_main_sheet[f"C{tenant_row}"] = updated_tenant_age
                        tenant_main_sheet[f"D{tenant_row}"] = updated_tenant_occu
                        tenant_main_sheet[f"E{tenant_row}"] = updated_tenant_gender
                        tenant_main_sheet[f"F{tenant_row}"] = updated_tenant_phone
                        tenant_main_sheet[f"G{tenant_row}"] = updated_tenant_email
                        if tenant_hd:
                            payment_main_sheet[f"B{d_r}"] = updated_tenant_name
                        excel_connection.save(excel_file)
                        messagebox.showinfo("SUCCESS", "Data is sucessfully updatad.")
                        manage_updete_tl_tenants.destroy()
                        tenant()
                else:
                    messagebox.showerror("ERROR", f"Action can't be accomplish.\nUnit {new_tenant_unit} is already occupied." )

        tenant_updete_tl_main_frame = Frame(manage_updete_tl_tenants)
        tenant_updete_tl_main_frame.pack(padx=(10, 5), pady=(0, 10))

        tenant_updete_tl_top_frame = Frame(tenant_updete_tl_main_frame)
        tenant_updete_tl_title = Label(tenant_updete_tl_top_frame, 
                                    text=f"Update Unit {to_update_unit_num}", 
                                    font=("Arial Black", 30),
                        fg="#217346")
        tenant_updete_tl_top_frame.pack()
        tenant_updete_tl_title.pack(expand=TRUE)

        tenant_updete_tl_rigth_frame = Frame(tenant_updete_tl_main_frame)
        tenant_updete_tl_rigth_frame.pack(expand=TRUE, fill=BOTH, side=LEFT, padx=(0, 5))
        # Add new tenant
        tenant_updete_tl_add_new_frame = ttk.LabelFrame(tenant_updete_tl_rigth_frame, text="TENANT'S INFORMATION")
        tenant_updete_tl_add_new_frame.pack(fill=X, pady=(0, 5))
        # Unit number their going to occupy
        tenant_updete_tl_new_unit_number_label = ttk.Label(tenant_updete_tl_add_new_frame, 
                                                        text="Unit #:", font=("", 13, "bold"))
        tenant_updete_tl_new_unit_number_enrty = ttk.Entry(tenant_updete_tl_add_new_frame, font=("", 13, "bold"), width=10)
        tenant_updete_tl_new_unit_number_label.grid(row=0, column=0, sticky=W, padx=20)
        tenant_updete_tl_new_unit_number_enrty.grid(row=0, column=1, sticky=W, pady=(0, 10))
        # Tenant Info
        def tenant_retrieve_old_name():
            if tenant_update_tl_name_var.get() == 0:
                tenant_updete_tl_new_name_entry.delete(0, END)
            else:
                tenant_updete_tl_new_name_entry.delete(0, END)
                tenant_updete_tl_new_name_entry.insert(0, update_tl_tenant_l[1])
        tenant_updete_tl_new_name_label = ttk.Label(tenant_updete_tl_add_new_frame, 
                                                        text="Name:", font=("", 13, "bold"))
        tenant_updete_tl_new_name_entry = ttk.Entry(tenant_updete_tl_add_new_frame)
        tenant_update_tl_name_var = IntVar()
        tenant_update_tl_name_check = ttk.Checkbutton(tenant_updete_tl_add_new_frame,
                                                        text="Retrieve old data", 
                                                        variable=tenant_update_tl_name_var,
                                                        command=tenant_retrieve_old_name)
        tenant_updete_tl_new_name_label.grid(row=2, column=0, sticky=W, padx=20)
        tenant_updete_tl_new_name_entry.grid(row=2, column=1, sticky=W)
        tenant_update_tl_name_check.grid(row=3, columnspan=2, sticky=W, padx=40, pady=(0, 10))

        
        def tenant_retrieve_old_age():
            if tenant_update_tl_age_var.get() == 0:
                tenant_updete_tl_new_age_entry.delete(0, END)
            else:
                tenant_updete_tl_new_age_entry.delete(0, END)
                tenant_updete_tl_new_age_entry.insert(0, update_tl_tenant_l[2])
        tenant_updete_tl_new_age_label = ttk.Label(tenant_updete_tl_add_new_frame, 
                                                        text="Age:", font=("", 13, "bold"))
        tenant_updete_tl_new_age_entry = ttk.Spinbox(tenant_updete_tl_add_new_frame, 
                                                    from_=1, to=100,
                                                    width=15)
        tenant_update_tl_age_var = IntVar()
        tenant_update_tl_age_check = ttk.Checkbutton(tenant_updete_tl_add_new_frame,
                                                        text="Retrieve old data", 
                                                        variable=tenant_update_tl_age_var,
                                                        command=tenant_retrieve_old_age)
        tenant_updete_tl_new_age_label.grid(row=4, column=0, sticky=W, padx=20)
        tenant_updete_tl_new_age_entry.grid(row=4, column=1, sticky=W)
        tenant_update_tl_age_check.grid(row=5, columnspan=2, sticky=W, padx=40, pady=(0, 10))


        def tenant_retrieve_old_occupa():
            if tenant_update_tl_occupation_var.get() == 0:
                tenant_updete_tl_new_occupation_entry.delete(0, END)
            else:
                tenant_updete_tl_new_occupation_entry.delete(0, END)
                tenant_updete_tl_new_occupation_entry.insert(0, update_tl_tenant_l[3])
        tenant_updete_tl_new_occupation_label = ttk.Label(tenant_updete_tl_add_new_frame, 
                                                        text="Occupation:",
                                                        font=("", 13, "bold"))
        tenant_updete_tl_new_occupation_entry = ttk.Entry(tenant_updete_tl_add_new_frame)
        tenant_update_tl_occupation_var = IntVar()
        tenant_update_tl_occupation_check = ttk.Checkbutton(tenant_updete_tl_add_new_frame,
                                                        text="Retrieve old data", 
                                                        variable=tenant_update_tl_occupation_var,
                                                        command=tenant_retrieve_old_occupa)
        tenant_updete_tl_new_occupation_label.grid(row=6, column=0, sticky=W, padx=20)
        tenant_updete_tl_new_occupation_entry.grid(row=6, column=1, sticky=W)
        tenant_update_tl_occupation_check.grid(row=7, columnspan=2, sticky=W, padx=40, pady=(0, 10))

        
        def tenant_retrieve_old_gender():
            if tenant_update_tl_gender_var.get() == 0:
                tenant_updete_tl_new_gender_var.set("")
            else:
                tenant_updete_tl_new_gender_var.set("")
                tenant_updete_tl_new_gender_var.set(update_tl_tenant_l[4])
        tenant_updete_tl_new_gender_frame = Frame(tenant_updete_tl_add_new_frame)
        tenant_updete_tl_new_gender_var = StringVar()
        tenant_updete_tl_new_gender_label = ttk.Label(tenant_updete_tl_new_gender_frame, 
                                                    text="Gender:", font=("", 13, "bold"))
        tenant_updete_tl_male_btn = ttk.Radiobutton(tenant_updete_tl_new_gender_frame, 
                                                    text="MALE", variable=tenant_updete_tl_new_gender_var, value="Male")
        tenant_updete_tl_female_btn = ttk.Radiobutton(tenant_updete_tl_new_gender_frame, 
                                                    text="FEMALE", variable=tenant_updete_tl_new_gender_var, 
                                                    value="Female")
        tenant_updete_tl_other_btn = ttk.Radiobutton(tenant_updete_tl_new_gender_frame, 
                                                    text="OTHER", variable=tenant_updete_tl_new_gender_var, value="Other")
        tenant_update_tl_gender_var = IntVar()
        tenant_update_tl_gender_check = ttk.Checkbutton(tenant_updete_tl_add_new_frame,
                                                        text="Retrieve old data", 
                                                        variable=tenant_update_tl_gender_var,
                                                        command=tenant_retrieve_old_gender)
        tenant_updete_tl_new_gender_frame.grid(row=8, columnspan=2, padx=13)
        tenant_update_tl_gender_check.grid(row=9, columnspan=2, sticky=W, padx=40, pady=(0, 10))
        tenant_updete_tl_new_gender_label.grid(row=0, column=0, padx=(7, 20))
        tenant_updete_tl_male_btn.grid(row=0, column=1, padx=20)
        tenant_updete_tl_female_btn.grid(row=0, column=2, padx=(0, 20))
        tenant_updete_tl_other_btn.grid(row=0, column=3,padx=(0, 5))


        def tenant_retrieve_old_phone():
            if tenant_update_tl_phone_number_var.get() == 0:
                tenant_updete_tl_phone_number.delete(0, END)
            else:
                tenant_updete_tl_phone_number.delete(0, END)
                tenant_updete_tl_phone_number.insert(0, update_tl_tenant_l[5])
        tenant_updete_tl_new_phone_number_label = ttk.Label(tenant_updete_tl_add_new_frame, 
                                                    text="Phone #:", font=("", 13, "bold"))
        tenant_updete_tl_phone_number = ttk.Entry(tenant_updete_tl_add_new_frame)
        tenant_update_tl_phone_number_var = IntVar()
        tenant_update_tl_phone_number_check = ttk.Checkbutton(tenant_updete_tl_add_new_frame,
                                                        text="Retrieve old data", 
                                                        variable=tenant_update_tl_phone_number_var,
                                                        command=tenant_retrieve_old_phone)
        tenant_updete_tl_new_phone_number_label.grid(row=10, column=0, sticky=W, padx=20)
        tenant_updete_tl_phone_number.grid(row=10, column=1, sticky=W)
        tenant_update_tl_phone_number_check.grid(row=11, columnspan=2, sticky=W, padx=40, pady=(0, 10))

        
        def tenant_retrieve_old_email():
            if tenant_update_tl_email_var.get() == 0:
                tenant_updete_tl_email_entry.delete(0, END)
            else:
                tenant_updete_tl_email_entry.delete(0, END)
                tenant_updete_tl_email_entry.insert(0, update_tl_tenant_l[6])
        tenant_updete_tl_new_email_label = ttk.Label(tenant_updete_tl_add_new_frame, 
                                                    text="Email:", font=("", 13, "bold"))
        tenant_updete_tl_email_entry = ttk.Entry(tenant_updete_tl_add_new_frame)
        tenant_update_tl_email_var = IntVar()
        tenant_update_tl_email_check = ttk.Checkbutton(tenant_updete_tl_add_new_frame,
                                                        text="Retrieve old data", 
                                                        variable=tenant_update_tl_email_var,
                                                        command=tenant_retrieve_old_email)
        tenant_updete_tl_new_email_label.grid(row=12, column=0, sticky=W, padx=20)
        tenant_updete_tl_email_entry.grid(row=12, column=1, sticky=W)
        tenant_update_tl_email_check.grid(row=13, columnspan=2, sticky=W, padx=40, pady=(0, 10))

        tenant_updete_tl_add_btn = Button(tenant_updete_tl_add_new_frame, 
                                        text="UPDATE", relief='flat', font=("", 13, "bold"), 
                                        bg='#217346', command=to_update_tenant)
        tenant_updete_tl_add_btn.grid(row=14, columnspan=2, sticky="we", padx=5, pady=(0, 5))

        exit_updete_tl_tenant_management = ttk.Button(tenant_updete_tl_rigth_frame, 
                                                    text="TENANTS MANAGEMENT",
                                                    command=update_tl_exit)
        exit_updete_tl_tenant_management.pack(fill=X, pady=(5, 0), ipady=3)

        tenant_updete_tl_new_unit_number_enrty.insert(0, to_update_unit_num)
        tenant_updete_tl_new_unit_number_enrty["state"] = "disable"
        
        manage_updete_tl_tenants.mainloop()


    def tenant_to_delete(tenant_unit, verify):
        number_row_apart = 0
        for every_unit_num in range(2, apart_sheet.max_row+1):
            if tenant_unit == apart_sheet["A"+str(every_unit_num)].value:
                number_row_apart = every_unit_num
                break
        number_row = 1
        for every_unit_occupied in range(2, tenant_main_sheet.max_row+1):
            if tenant_unit == tenant_main_sheet["A"+str(every_unit_occupied)].value:
                number_row = every_unit_occupied
                break
        unit_pay = False
        p_r = 0
        for each_due in range(2, payment_main_sheet.max_row+1):
            if payment_main_sheet["A"+str(each_due)].value == tenant_unit:
                unit_pay = True
                p_r = each_due
        unit_n = tenant_unit
        rec_row = 0
        found = False
        for i in payment_record_sheet["A"]:
            rec_row += 1
            if unit_n == i.value:
                found = True
                break
        tenant_name_deleting = tenant_main_sheet[f"B{number_row}"].value
        is_deleting_tanant_data = messagebox.askquestion("Verifying", f"You're about to delete data of {tenant_name_deleting} in record? \
                                                \nNote: ACTION WILL DELETE ALL DATA ABOUT {tenant_name_deleting} PERMANENTLY \
                                                \n\nCLICK \"yes\" TO CONTINUE")
        if is_deleting_tanant_data == "yes":
            if found:
                payment_record_sheet.delete_rows(rec_row)
            if unit_pay:
                payment_main_sheet.delete_rows(p_r)
            for every_apart in range(2, main_sheet.max_row+1):
                if tenant_unit == main_sheet['A'+str(every_apart)].value:
                    unit_main_row = every_apart
                    break
            main_sheet["C"+str(unit_main_row)] = "NONE"
            main_sheet["B"+str(unit_main_row)] = "Vacant"
            tenant_main_sheet.delete_rows(number_row)
            apart_sheet[f"C{number_row_apart}"] = "Vacant"
            excel_connection.save(excel_file)
            messagebox.showinfo("SUCCESS", f"Data {tenant_name_deleting} is successfully deleted.")
            verify.destroy()
            tenant()
        else:
            tenant()


    def manage_apart_func():
        manage_tenants.destroy()
        apartment()


    def manage_payment_func_tent():
        manage_tenants.destroy()
        payment()


    def tenant_main_exit():
        manage_tenants.destroy()
        main_interface()


    def add_tenant():
        tenant_unit_number = tenant_new_unit_number_enrty.get().strip()
        tenant_name = tenant_new_name_entry.get().strip()
        tenant_age = tenant_new_age_entry.get().strip()
        tenant_occupation = tenant_new_occupation_entry.get().strip()
        tenant_gender = tenant_new_gender_var.get().strip()
        tenant_contact = tenant_phone_number.get().strip()
        tenant_email = tenant_email_entry.get().strip()

        tenant_new_info = [
            tenant_unit_number,
            tenant_name, 
            tenant_age,
            tenant_occupation,
            tenant_gender,
            tenant_contact,
            tenant_email
        ]
        if tenant_new_info[0] == "Unit #" or tenant_new_info[0] == "" or\
            tenant_new_info[1] == "Tenant's Name" or tenant_new_info[1] == "" or\
            tenant_new_info[2] == "Age" or tenant_new_info[2] == "" or\
            tenant_new_info[3] == "Occupation" or tenant_new_info[3] == "" or\
            tenant_new_info[4] == "" or\
            tenant_new_info[5] == "Contact Number" or tenant_new_info[4] == "" or\
            tenant_new_info[6] == "Email" or tenant_new_info[6] == "":
            messagebox.showerror("ERROR", f"MUST FILL ALL THE FIELDS!")
        else:
            found_apartment = is_unit_exist(tenant_unit_number)
            if found_apartment:
                for every_unit_num in range(2, apart_sheet.max_row+1):
                    if tenant_unit_number == apart_sheet["A"+str(every_unit_num)].value:
                        number_row = every_unit_num
                        break
                unit_has_tenant = False

                for every_tanant_data in range(2, tenant_main_sheet.max_row+1):
                    if tenant_unit_number == tenant_main_sheet["A"+str(every_tanant_data)].value:
                        unit_has_tenant = True
                        break
                    else:
                        unit_has_tenant = False
                if unit_has_tenant:
                    messagebox.showerror("ERROR", f"UNIT \"{tenant_unit_number}\" IS ALREADY OCCUPIED!")
                elif unit_has_tenant == False :
                    new_tenant_data = f"""
    New Tenant Data

    Unit Number: {tenant_new_info[0]}
    Name: {tenant_new_info[1]}
    Age: {tenant_new_info[2]}
    Occupation: {tenant_new_info[3]}
    Gendar: {tenant_new_info[4]}
    Contact Number: {tenant_new_info[5]}
    Email: {tenant_new_info[6]}

    click "Yes" to continue.
    """
                    add_new_tenant_ans = messagebox.askquestion("Verifying", new_tenant_data, icon="question")
                    if add_new_tenant_ans == "yes":
                        for every_apart in range(2, main_sheet.max_row+1):
                            if tenant_unit_number == main_sheet['A'+str(every_apart)].value:
                                unit_main_row = every_apart
                                break
                        main_sheet["C"+str(unit_main_row)] = tenant_name
                        main_sheet["B"+str(unit_main_row)] = "Occupied"
                        tenant_main_sheet.append(tenant_new_info)
                        apart_sheet[f"C{number_row}"] = "Occupied"
                        tenant_refresh_tree(tenant_tree)
                        excel_connection.save(excel_file)
                        tenant_refresh_entries_func()
                        messagebox.showinfo("Success", "Tenant Successfully Added.")
                        messagebox.showinfo("Set New Due Date", f"Set Due Date for the New Tenant {tenant_new_info[1]}")
                        manage_tenants.destroy()
                        payment()
            else:
                messagebox.showerror("ERROR", f"Unit {tenant_unit_number} does not exist.")


    def tenant_refresh_tree(ref_tenant_tree):
        apart_updated_data = list(tenant_main_sheet.values)

        ref_tenant_tree.delete(*ref_tenant_tree.get_children())
        for every_row_apart in apart_updated_data[1:]:
            ref_tenant_tree.insert("", END, values=every_row_apart)


    def tenant_refresh_entries_func():
        tenant_new_unit_number_enrty.delete(0, END)
        tenant_new_unit_number_enrty.insert(0, "Unit #")

        tenant_new_name_entry.delete(0, END)
        tenant_new_name_entry.insert(0, "Tenant's Name")

        tenant_new_age_entry.delete(0, END)
        tenant_new_age_entry.insert(0, "Age")

        tenant_new_occupation_entry.delete(0, END)
        tenant_new_occupation_entry.insert(0, "Occupation")

        tenant_new_gender_var.set("")

        tenant_phone_number.delete(0, END)
        tenant_phone_number.insert(0, "Contact Number")

        tenant_email_entry.delete(0, END)
        tenant_email_entry.insert(0, "Email")


    tenant_main_frame = Frame(manage_tenants)
    tenant_main_frame.pack(padx=10, pady=(0, 10), expand=TRUE)

    tenant_top_frame = Frame(tenant_main_frame)
    tenant_title = Label(tenant_top_frame, text="Tenants Management", font=("Arial Black", 70),
                    fg="#217346")
    tenant_top_frame.pack()
    tenant_title.pack(expand=TRUE)

    tenant_rigth_frame = Frame(tenant_main_frame)
    tenant_search_frame = ttk.LabelFrame(tenant_rigth_frame, text="SEARCH • UPDATE • DELETE")
    tenant_search_entry = ttk.Entry(tenant_search_frame, font=("", 13, "bold"), width=10)
    tenant_search_entry.insert(0, "Unit #")
    def tenant_search_entry_i(event):
        if tenant_search_entry.get().strip() == "" or \
            tenant_search_entry.get().strip() == "Unit #":
            tenant_search_entry.delete(0, END)
    def tenant_search_entry_o(event):
        if tenant_search_entry.get().strip() == "":
            tenant_search_entry.insert(0, "Unit #")
    tenant_search_entry.bind("<FocusIn>", tenant_search_entry_i)
    tenant_search_entry.bind("<FocusOut>", tenant_search_entry_o)
    tenant_search_btn = Button(tenant_search_frame, text="SEARCH", relief='flat', font=("", 13, "bold"), bg='#217346', command=tenant_to_search)
    tenant_update_btn = Button(tenant_search_frame, text="UPDATE", relief='flat', font=("", 13, "bold"), bg='#217346', 
                               command=lambda:verify(management=manage_tenants, unit_to_update=tenant_search_entry.get().strip(), 
                                                     action="updating_tanant", function_for_action=tenant_to_update))
    tenant_delete_btn = ttk.Button(tenant_search_frame, text="DELETE", 
                                   command=lambda:verify(management=manage_tenants, 
                                                         action="deleting_tenant", unit_to_delete=tenant_search_entry.get().strip(), 
                                                         function_for_action=tenant_to_delete))
    tenant_rigth_frame.pack(expand=TRUE, fill=BOTH, side=LEFT, padx=(0, 5))
    tenant_search_frame.pack(expand=TRUE, fill=BOTH, pady=(0, 10))
    tenant_search_entry.pack(expand=TRUE, fill=BOTH, padx=5, pady=5)
    tenant_search_btn.pack(side=LEFT, expand=TRUE, fill=BOTH, padx=5, pady=(0, 5))
    tenant_update_btn.pack(side=LEFT, expand=TRUE, fill=BOTH, padx=0, pady=(0, 5))
    tenant_delete_btn.pack(side=LEFT, expand=TRUE, fill=BOTH, padx=5, pady=(0, 5))

    # Add new tenant
    tenant_add_new_frame = ttk.LabelFrame(tenant_rigth_frame, text="ADD NEW TENANT")
    tenant_add_new_frame.pack(fill=X, pady=(0, 5))
    # Unit number their going to occupy
    tenant_new_unit_number_enrty = ttk.Entry(tenant_add_new_frame, font=("", 13, "bold"))
    tenant_new_unit_number_enrty.insert(0, "Unit #")
    def tenant_unit_new_entry_i(event):
        if tenant_new_unit_number_enrty.get().strip() == "" or \
            tenant_new_unit_number_enrty.get().strip() == "Unit #":
            tenant_new_unit_number_enrty.delete(0, END)
    def tenant_unit_new_entry_o(event):
        if tenant_new_unit_number_enrty.get().strip() == "":
            tenant_new_unit_number_enrty.insert(0, "Unit #")
    tenant_new_unit_number_enrty.bind("<FocusIn>", tenant_unit_new_entry_i)
    tenant_new_unit_number_enrty.bind("<FocusOut>", tenant_unit_new_entry_o)
    tenant_new_unit_number_enrty.pack(fill=X, padx=5, pady=5)

    # Tenant Info
    tenant_new_name_entry = ttk.Entry(tenant_add_new_frame, font=("", 13, "bold"))
    tenant_new_name_entry.insert(0, "Tenant's Name")
    def tenant_name_new_entry_i(event):
        if tenant_new_name_entry.get().strip() == "" or \
            tenant_new_name_entry.get().strip() == "Tenant's Name":
            tenant_new_name_entry.delete(0, END)
    def tenant_name_new_entry_o(event):
        if tenant_new_name_entry.get().strip() == "":
            tenant_new_name_entry.insert(0, "Tenant's Name")
    tenant_new_name_entry.bind("<FocusIn>", tenant_name_new_entry_i)
    tenant_new_name_entry.bind("<FocusOut>", tenant_name_new_entry_o)
    tenant_new_name_entry.pack(fill=X, padx=5, pady=(0, 5))

    tenant_new_age_entry = ttk.Spinbox(tenant_add_new_frame, from_=1, to=100, font=("", 13, "bold"))
    tenant_new_age_entry.insert(0, "Age")
    def tenant_name_age_entry_i(event):
        if tenant_new_age_entry.get().strip() == "" or \
            tenant_new_age_entry.get().strip() == "Age":
            tenant_new_age_entry.delete(0, END)
    def tenant_name_age_entry_o(event):
        if tenant_new_age_entry.get().strip() == "":
            tenant_new_age_entry.insert(0, "Age")
    tenant_new_age_entry.bind("<FocusIn>", tenant_name_age_entry_i)
    tenant_new_age_entry.bind("<FocusOut>", tenant_name_age_entry_o)
    tenant_new_age_entry.pack(fill=X, padx=5, pady=(0, 5))

    tenant_new_occupation_entry = ttk.Entry(tenant_add_new_frame, font=("", 13, "bold"))
    tenant_new_occupation_entry.insert(0, "Occupation")
    def tenant_occupation_new_entry_i(event):
        if tenant_new_occupation_entry.get().strip() == "" or \
            tenant_new_occupation_entry.get().strip() == "Occupation":
            tenant_new_occupation_entry.delete(0, END)
    def tenant_occupation_new_entry_o(event):
        if tenant_new_occupation_entry.get().strip() == "":
            tenant_new_occupation_entry.insert(0, "Occupation")
    tenant_new_occupation_entry.bind("<FocusIn>", tenant_occupation_new_entry_i)
    tenant_new_occupation_entry.bind("<FocusOut>", tenant_occupation_new_entry_o)
    tenant_new_occupation_entry.pack(fill=X, padx=5, pady=(0, 5))

    tenant_new_gender_frame = Frame(tenant_add_new_frame)
    tenant_new_gender_var = StringVar()
    tenant_new_gender_label = ttk.Label(tenant_new_gender_frame, text="Gender:", font=("", 13, "bold"))
    tenant_male_btn = ttk.Radiobutton(tenant_new_gender_frame, text="MALE", variable=tenant_new_gender_var, value="Male")
    tenant_female_btn = ttk.Radiobutton(tenant_new_gender_frame, text="FEMALE", variable=tenant_new_gender_var, value="Female")
    tenant_other_btn = ttk.Radiobutton(tenant_new_gender_frame, text="OTHER", variable=tenant_new_gender_var, value="Other")
    tenant_new_gender_frame.pack(fill=X, padx=5, pady=(0, 5))
    tenant_new_gender_label.grid(row=0, column=0, padx=(7, 20))
    tenant_male_btn.grid(row=0, column=1, padx=20)
    tenant_female_btn.grid(row=0, column=2, padx=(0, 20))
    tenant_other_btn.grid(row=0, column=3)

    tenant_phone_number = ttk.Entry(tenant_add_new_frame, font=("", 13, "bold"))
    tenant_phone_number.insert(0, "Contact Number")
    def tenant_phone_number_new_entry_i(event):
        if tenant_phone_number.get().strip() == "" or \
            tenant_phone_number.get().strip() == "Contact Number":
            tenant_phone_number.delete(0, END)
    def tenant_phone_number_new_entry_o(event):
        if tenant_phone_number.get().strip() == "":
            tenant_phone_number.insert(0, "Contact Number")
    tenant_phone_number.bind("<FocusIn>", tenant_phone_number_new_entry_i)
    tenant_phone_number.bind("<FocusOut>", tenant_phone_number_new_entry_o)
    tenant_phone_number.pack(fill=X, padx=5, pady=(0, 5))

    tenant_email_entry = ttk.Entry(tenant_add_new_frame, font=("", 13, "bold"))
    tenant_email_entry.insert(0, "Email")
    def tenant_email_new_entry_i(event):
        if tenant_email_entry.get().strip() == "" or \
            tenant_email_entry.get().strip() == "Email":
            tenant_email_entry.delete(0, END)
    def tenant_email_new_entry_o(event):
        if tenant_email_entry.get().strip() == "":
            tenant_email_entry.insert(0, "Email")
    tenant_email_entry.bind("<FocusIn>", tenant_email_new_entry_i)
    tenant_email_entry.bind("<FocusOut>", tenant_email_new_entry_o)
    tenant_email_entry.pack(fill=X, padx=5, pady=(0, 5))

    tenant_add_btn = Button(tenant_add_new_frame, text="ADD TENANT", relief='flat', font=("", 13, "bold"), 
                            bg='#217346', command=add_tenant)
    tenant_add_btn.pack(fill=X, padx=5, pady=(0, 5))

    tenant_other_management_frame = Frame(tenant_rigth_frame)
    tenant_apart_man_btn = ttk.Button(tenant_other_management_frame, text="MANAGE APARTMENT", command=manage_apart_func)
    tenant_paym_btn = ttk.Button(tenant_other_management_frame, text="MANAGE PAYMENTS", command=manage_payment_func_tent)
    tenant_other_management_frame.pack(fill=X)
    tenant_apart_man_btn.pack(side=LEFT, expand=TRUE, fill=X, padx=(0, 5), ipadx=10, ipady=3)
    tenant_paym_btn.pack(side=LEFT, expand=TRUE, fill=X, ipadx=10, ipady=3)

    tenant_left_frame = Frame(tenant_main_frame)
    tenant_left_top_frame = Frame(tenant_left_frame)
    excel_items = list(tenant_main_sheet.values)
    tenant_tree = ttk.Treeview(tenant_left_top_frame, show='headings', columns=excel_items[0])
    tenant_scroll_bar_y = ttk.Scrollbar(tenant_left_top_frame, orient='vertical', command=tenant_tree.yview)
    tenant_tree.configure(yscrollcommand=tenant_scroll_bar_y.set)
    for heading in excel_items[0]:
        if heading == "UNIT":
            tenant_tree.column(heading, anchor=CENTER, width=40)
            tenant_tree.heading(heading, text=heading)
        elif heading == "GENDER":
            tenant_tree.column(heading, anchor=CENTER, width=70)
            tenant_tree.heading(heading, text=heading)
        elif heading == "AGE":
            tenant_tree.column(heading, anchor=CENTER, width=50)
            tenant_tree.heading(heading, text=heading)
        elif heading == "CONTACT #":
            tenant_tree.column(heading, anchor=CENTER, width=150)
            tenant_tree.heading(heading, text=heading)
        elif heading == "OCCUPATION":
            tenant_tree.column(heading, anchor=CENTER, width=100)
            tenant_tree.heading(heading, text=heading)
        else:
            tenant_tree.column(heading, anchor=CENTER)
            tenant_tree.heading(heading, text=heading)
    for value in excel_items[1:]:
        tenant_tree.insert('', END, values=value)
    tenant_left_top_frame.pack(expand=True, fill=BOTH)
    tenant_left_frame.pack(expand=TRUE, fill=BOTH, pady=(7, 0))
    tenant_tree.pack(side=LEFT, expand=TRUE, fill=Y)
    tenant_scroll_bar_y.pack(side=LEFT, fill=Y)
    exit_tenant_management = ttk.Button(tenant_left_frame, text="MAIN PAGE", command=tenant_main_exit)
    exit_tenant_management.pack(fill=X, pady=(5, 0), ipady=3)

    manage_tenants.mainloop()


def payment():
    manage_payments = tk.Tk()
    manage_payments.title("Payment Management")
    manage_payments.resizable(0, 0)

    # Theme
    style = ttk.Style(manage_payments)
    manage_payments.tk.call("source", 'theme_forest-dark.tcl')
    style.theme_use("forest-dark")

    def search_history_payment():
        unit_n = payment_search_history_entry.get().strip()
        if unit_n != "Unit #" and unit_n != "":
            row = 0
            found = False
            for i in payment_record_sheet["A"]:
                row += 1
                if unit_n == i.value:
                    found = True
                    break
            if found:
                unit_history = list(payment_record_sheet[row])
                history = Toplevel()
                history.title("History")
                history.resizable(0, 0)
                history_main_frame = Frame(history)
                history_main_frame.pack(expand=True, padx=5, pady=5)
                history_tree = ttk.Treeview(history_main_frame, show='headings', columns="1")
                history_scroll_bar_y = ttk.Scrollbar(history_main_frame, orient='vertical', command=history_tree.yview)
                history_tree.column("1", anchor=CENTER, width=600)
                history_tree.heading("1", text=f"History for Unit # \"{unit_n}\"")
                list_of_record = [] 
                for every_record in unit_history[1:]:
                    if every_record.value != None:
                        list_of_record.append([every_record.value])
                for each_record in list_of_record:
                    history_tree.insert("", END, values=each_record)
                history_tree.pack(side=LEFT, expand=True, fill=BOTH)
                history_scroll_bar_y.pack(side=LEFT, fill=Y)
            else:
                messagebox.showerror("ERROR", f"Can't find payment record for \"{unit_n}\".")
        else:
            messagebox.showerror('ERROR', "Must provide a Unit Number.")


    def deleting_history_payment(unit_n, verify):
        row = 0
        for i in payment_record_sheet["A"]:
            row += 1
            if unit_n == i.value:
                break
        is_agree = messagebox.askquestion("Are you sure?", f"Are you sure you want to delete payment history for Unit {unit_n}?\n\nClick \"Yes\" to continue") 
        if is_agree == "yes":
            unit_number = payment_record_sheet["A"+str(row)].value
            updated = [unit_number]
            payment_record_sheet.delete_rows(row)
            payment_record_sheet.append(updated)
            excel_connection.save(excel_file)
            verify.destroy()
            payment()
        else:
            verify.destroy()
            payment()


    def adding_to_record(unit_number, details):
        unit_n = unit_number
        row = 0
        found = False
        for i in payment_record_sheet["A"]:
            row += 1
            if unit_n == i.value:
                found = True
                break

        if found:
            current_l = [unit_n, details]
            current_t = list(payment_record_sheet[row])
            for item in current_t:
                if item.value != unit_n:
                    current_l.append(item.value)

            payment_record_sheet.delete_rows(row)
            payment_record_sheet.append(current_l)
            excel_connection.save(excel_file)


    def make_payment(mp_unit_num, verify, mp_day, mp_month, mp_year, mp_for, amount, mp_name):
        date = f"{mp_day} / {mp_month} / {mp_year}"
        mp_amount = int(amount)
        unit_r = 0
        for every_d in range(2, payment_main_sheet.max_row+1):
            if mp_unit_num == payment_main_sheet[f"A{every_d}"].value:
                unit_r = every_d
                break
        curent_rent_bal = int(payment_main_sheet[f"C{unit_r}"].value)
        curent_rent_stat = payment_main_sheet[f"D{unit_r}"].value
        curent_uti_bal = int(payment_main_sheet[f"E{unit_r}"].value)
        curent_uti_stat =  payment_main_sheet[f"F{unit_r}"].value

        # New amount initial value
        rent_new_bal = curent_rent_bal
        rent_new_status =  curent_rent_stat
        utilities_new_bal = curent_uti_bal
        utilities_new_status = curent_uti_stat

        if mp_for == "Rent":
            rent_new_bal = curent_rent_bal - mp_amount
            if rent_new_bal <= 0:
                rent_new_status = "PAID"
            else:
                rent_new_status = curent_rent_stat
        elif mp_for == "Utilities":
            utilities_new_bal = curent_uti_bal - mp_amount
            if utilities_new_bal <= 0:
                utilities_new_status = "PAID"
            else:
                utilities_new_status = curent_uti_stat
        elif mp_for == "Rent and Utilities":
            if mp_amount > curent_rent_bal:
                mp_amount -= curent_rent_bal
                if mp_amount >= curent_uti_bal:
                    mp_amount -= curent_uti_bal
                    utilities_new_bal = 0
                    rent_new_bal = -abs(mp_amount)
                elif mp_amount >= 1 and mp_amount < curent_uti_bal:
                    utilities_new_bal = curent_uti_bal - mp_amount
                    rent_new_bal = 0
            elif mp_amount <= curent_rent_bal:
                rent_new_bal = curent_rent_bal - mp_amount
                utilities_new_bal = curent_uti_bal

            if utilities_new_bal <= 0 and rent_new_bal <= 0:
                rent_new_status = "PAID"
                utilities_new_status = "PAID"
            elif utilities_new_bal == 0 and rent_new_bal >= 1:
                utilities_new_status = "PAID"
            elif rent_new_bal <= 0:
                rent_new_status = "PAID"

        is_payment_confirm = messagebox.askquestion("Verifying", f"You are about to make a payment for {mp_for.title()} for Unti {mp_unit_num}?\n\nClick \"Yes\" continue.")
        if is_payment_confirm == "yes":
            t = f"{mp_name}, {amount}, {mp_for}, {date}"
            adding_to_record(mp_unit_num, t)

            payment_main_sheet[f"C{unit_r}"] = rent_new_bal
            payment_main_sheet[f"D{unit_r}"] = rent_new_status
            payment_main_sheet[f"E{unit_r}"] = utilities_new_bal
            payment_main_sheet[f"F{unit_r}"] = utilities_new_status
            if rent_new_status == "PAID" and utilities_new_status == "PAID":
                payment_main_sheet[f"G{unit_r}"] = "NONE"
            excel_connection.save(excel_file)
            verify.destroy()
            payment()
        else:
            curent_rent_bal = int(payment_main_sheet[f"C{unit_r}"].value)
            curent_rent_stat = payment_main_sheet[f"D{unit_r}"].value
            curent_uti_bal = int(payment_main_sheet[f"E{unit_r}"].value)
            curent_uti_stat =  payment_main_sheet[f"F{unit_r}"].value
            verify.destroy()
            payment()


    def payment_apart():
        manage_payments.destroy()
        apartment()


    def payment_tenant():
        manage_payments.destroy()
        tenant()


    def payment_exit():
        manage_payments.destroy()
        main_interface()


    def payment_set_ndd(ndd_unit_number, _verify, set_ndd_day, set_ndd_month, set_ndd_year):
        rent_status = "UNPAID"
        nnd_utilities_status = "UNPAID"
        due_date = f"{set_ndd_day} / {set_ndd_month} / {set_ndd_year}" 
        apart_unit_row = 0
        apart_unit_found = False
        for apart_every_unit in range(2, apart_sheet.max_row+1):
            if ndd_unit_number == apart_sheet[f"A{apart_every_unit}"].value:
                apart_unit_row = apart_every_unit
                apart_unit_found = True
                break
        if apart_unit_found:
            apart_unit_rent_rate = int(apart_sheet[f"B{apart_unit_row}"].value)
            current_utili_status_water = apart_sheet[f"N{apart_unit_row}"].value
            current_utili_status_elec = apart_sheet[f"P{apart_unit_row}"].value
            current_utili_status_inter = apart_sheet[f"R{apart_unit_row}"].value
            # Water stat
            if current_utili_status_water == "INCLUDED":
                fix_water_rate = int(apart_sheet[f"O{apart_unit_row}"].value)
            elif current_utili_status_water != "INCLUDED":
                fix_water_rate = 0

            # Electrcity stat
            if current_utili_status_elec == "INCLUDED":
                fix_elec_rate = int(apart_sheet[f"Q{apart_unit_row}"].value)
            elif current_utili_status_elec != "INCLUDED":
                fix_elec_rate = 0

            # Internet stat
            if current_utili_status_inter == "INCLUDED":
                fix_inter_rate = int(apart_sheet[f"S{apart_unit_row}"].value)
            elif current_utili_status_inter != "INCLUDED":
                fix_inter_rate = 0


            total_fix_utili = fix_water_rate + fix_elec_rate + fix_inter_rate

            tenant_unit_row = 0
            tenant_unit_found = False
            for tenant_every_unit in range(2, tenant_main_sheet.max_row+1):
                if ndd_unit_number == tenant_main_sheet[f"A{tenant_every_unit}"].value:
                    tenant_unit_found = True
                    tenant_unit_row = tenant_every_unit
                    break
            if tenant_unit_found:
                tenant_name = tenant_main_sheet[f"B{tenant_unit_row}"].value
                due_unit_row = 0
                due_unit_found = False
                for due_every_unit in range(2, payment_main_sheet.max_row+1):
                    if ndd_unit_number == payment_main_sheet[f"A{due_every_unit}"].value:
                        due_unit_row = due_every_unit
                        due_unit_found = True
                        break

                if ndd_unit_number != "Unit #" and ndd_unit_number != ""\
                and set_ndd_day != "Day" and set_ndd_day != ""\
                and set_ndd_month != "Month" and set_ndd_month != ""\
                and set_ndd_year != "Year" and set_ndd_year != "":
                    if due_unit_found:
                        current_unit_rent_bal = int(payment_main_sheet[f"C{due_unit_row}"].value)
                        current_rent_status = payment_main_sheet[f"D{due_unit_row}"].value
                        current_unit_utili_bal = int(payment_main_sheet[f"E{due_unit_row}"].value)
                        current_utili_status = payment_main_sheet[f"F{due_unit_row}"].value
                        
                        # Rent
                        if current_unit_rent_bal < 0:
                            apart_unit_rent_rate += current_unit_rent_bal
                        elif current_unit_rent_bal == 0:
                            apart_unit_rent_rate = int(apart_sheet[f"B{apart_unit_row}"].value)
                        elif current_unit_rent_bal > 0:
                            apart_unit_rent_rate += current_unit_rent_bal

                        # Utilities
                        if current_unit_utili_bal < 0:
                            total_fix_utili += current_unit_utili_bal
                        elif current_unit_utili_bal == 0:
                            total_fix_utili = fix_water_rate + fix_elec_rate + fix_inter_rate
                        elif current_unit_utili_bal > 0:
                            total_fix_utili += current_unit_utili_bal

                        if apart_unit_rent_rate <= 0:
                            rent_status = "PAID"
                        if total_fix_utili <= 0:
                            nnd_utilities_status = "PAID"

                        if rent_status == "PAID" and nnd_utilities_status == "PAID":
                            due_date = "NONE"
                        u_sure_add_set = messagebox.askquestion("Are you sure?", f"You're about to set new due date in the record for tenant{tenant_name.title()}\n\nClick \"Yes\" to continue.")
                        if u_sure_add_set == "yes":
                            set_due_date_list = [
                                apart_unit_rent_rate,
                                rent_status,
                                total_fix_utili,
                                nnd_utilities_status,
                                due_date
                            ]
                            payment_main_sheet[f"C{due_unit_row}"] = set_due_date_list[0]
                            payment_main_sheet[f"D{due_unit_row}"] = set_due_date_list[1]
                            payment_main_sheet[f"E{due_unit_row}"] = set_due_date_list[2]
                            payment_main_sheet[f"F{due_unit_row}"] = set_due_date_list[3]
                            payment_main_sheet[f"G{due_unit_row}"] = set_due_date_list[4]
                            excel_connection.save(excel_file)
                            messagebox.showinfo("SUCCESS", "You are successfully add new record.")
                            _verify.destroy()
                            payment()
                    else:
                        if total_fix_utili <= 0:
                            nnd_utilities_status = "PAID"
                        add_due_date_list = [
                            ndd_unit_number,
                            tenant_name, 
                            apart_unit_rent_rate,
                            rent_status,
                            total_fix_utili,
                            nnd_utilities_status,
                            due_date
                        ]
                        details = f"""
    You're about to add new due date in the 
    record for tenant{tenant_name.title()}

    Check the deatils:

    Unit Number: {ndd_unit_number}
    Tenant's Name: {tenant_name}
    Rent Rate: {apart_unit_rent_rate}.00
    Utilities Rate: {total_fix_utili}.00

    Click \"Yes\" to continue.
""" 
                        u_sure_add_new = messagebox.askquestion("Are you sure?", details)
                        if u_sure_add_new == "yes":
                            payment_main_sheet.append(add_due_date_list)
                            new_due = [ndd_unit_number]
                            payment_record_sheet.append(new_due)
                            excel_connection.save(excel_file)
                            messagebox.showinfo("SUCCESS", "You are successfully add new record.")
                            _verify.destroy()
                            payment()


    payment_main_frame = Frame(manage_payments)
    payment_main_frame.pack(expand=True, fill=BOTH, padx=10, pady=(0, 10))


    payment_title_frame = Frame(payment_main_frame)
    payment_entries_frame = Frame(payment_main_frame)
    payment_right_frame = Frame(payment_main_frame)
    payment_tree_frame = Frame(payment_right_frame)
    payment_title_frame.pack(side=TOP)
    payment_entries_frame.pack(side=LEFT, padx=(0, 5))
    payment_right_frame.pack(side=RIGHT, expand=TRUE, fill=BOTH)
    payment_tree_frame.pack(expand=TRUE, fill=BOTH)


    payment_management_title = Label(payment_title_frame, text="Payment Management", 
                                    font=("Arial Black", 65),fg="#217346")
    payment_management_title.pack(expand=TRUE, fill=BOTH)

    payment_history_frame = ttk.LabelFrame(payment_entries_frame, 
                                        text="Search • Delete (Payment History)")
    payment_search_history_entry = ttk.Entry(payment_history_frame, font=("", 13, "bold"), width=10)
    payment_search_history_entry.insert(0, "Unit #")
    def payment_search_entry_i(event):
        if payment_search_history_entry.get().strip() == "" or \
            payment_search_history_entry.get().strip() == "Unit #":
            payment_search_history_entry.delete(0, END)
    def payment_search_entry_o(event):
        if payment_search_history_entry.get().strip() == "":
            payment_search_history_entry.insert(0, "Unit #")
    payment_search_history_entry.bind("<FocusIn>", payment_search_entry_i)
    payment_search_history_entry.bind("<FocusOut>", payment_search_entry_o)

    payment_search_history_btn = Button(payment_history_frame, 
                                    text="HISTORY", relief='flat', 
                                    font=("", 13, "bold"), bg='#217346',
                                    command=search_history_payment)
    payment_delete_history_btn = ttk.Button(payment_history_frame, 
                                    text="DELETE",
                                    command=lambda:verify(management=manage_payments, 
                                                          action="deleting_payment_history", 
                                                          unit_to_delete=payment_search_history_entry.get().strip(),
                                                          function_for_action=deleting_history_payment))
    payment_history_frame.pack(fill=X)
    payment_search_history_entry.pack(side=TOP, fill=X, padx=5, pady=(5, 0))
    payment_search_history_btn.pack(side=LEFT, fill=X, expand=True, padx=5, pady=5  )
    payment_delete_history_btn.pack(side=RIGHT, fill=X, expand=True, padx=(0, 5), pady=5)

    # Make Payment
    payment_make_payment_frame = ttk.LabelFrame(payment_entries_frame, text="Make Payment")
    payment_make_payment_unit_entry = ttk.Entry(payment_make_payment_frame, 
                                                font=("", 13, "bold"), width=10)
    payment_make_payment_unit_entry.insert(0, "Unit #")
    def payment_make_pay_entry_i(event):
        if payment_make_payment_unit_entry.get().strip() == "" or \
            payment_make_payment_unit_entry.get().strip() == "Unit #":
            payment_make_payment_unit_entry.delete(0, END)
    def payment_make_pay_entry_o(event):
        if payment_make_payment_unit_entry.get().strip() == "":
            payment_make_payment_unit_entry.insert(0, "Unit #")
    payment_make_payment_unit_entry.bind("<FocusIn>", payment_make_pay_entry_i)
    payment_make_payment_unit_entry.bind("<FocusOut>", payment_make_pay_entry_o)

    payment_make_payment_amount_entry = ttk.Entry(payment_make_payment_frame, 
                                                font=("", 13, "bold"), width=10)
    payment_make_payment_amount_entry.insert(0, "Amount")
    def payment_make_amount_entry_i(event):
        if payment_make_payment_amount_entry.get().strip() == "" or \
            payment_make_payment_amount_entry.get().strip() == "Amount":
            payment_make_payment_amount_entry.delete(0, END)
    def payment_make_amount_entry_o(event):
        if payment_make_payment_amount_entry.get().strip() == "":
            payment_make_payment_amount_entry.insert(0, "Amount")
    payment_make_payment_amount_entry.bind("<FocusIn>", payment_make_amount_entry_i)
    payment_make_payment_amount_entry.bind("<FocusOut>", payment_make_amount_entry_o)

    payment_make_payment_name_entry = ttk.Entry(payment_make_payment_frame, font=("", 13, "bold"), width=10)
    payment_make_payment_name_entry.insert(0, "Name")
    def payment_make_name_entry_i(event):
        if payment_make_payment_name_entry.get().strip() == "" or \
            payment_make_payment_name_entry.get().strip() == "Name":
            payment_make_payment_name_entry.delete(0, END)
    def payment_make_name_entry_o(event):
        if payment_make_payment_name_entry.get().strip() == "":
            payment_make_payment_name_entry.insert(0, "Name")
    payment_make_payment_name_entry.bind("<FocusIn>", payment_make_name_entry_i)
    payment_make_payment_name_entry.bind("<FocusOut>", payment_make_name_entry_o)

    payment_make_payment_frame.pack(fill=X, pady=10)
    payment_make_payment_unit_entry.pack(fill=X, padx=5, pady=(5,0))
    payment_make_payment_amount_entry.pack(fill=X, padx=5, pady=5)
    payment_make_payment_name_entry.pack(fill=X, padx=5)

    what_to_pay = [
        "Rent",
        "Utilities",
        "Rent and Utilities"
    ]

    payment_make_pay_var = StringVar()
    payment_for_what_combo = ttk.Combobox(payment_make_payment_frame, values=what_to_pay, 
                                            textvariable=payment_make_pay_var, font=("", 13, "bold"))
    if payment_for_what_combo.get().strip() == "":
        payment_for_what_combo.insert(0, "PAYMENT FOR")
        payment_for_what_combo["state"] = "readonly"
    payment_for_what_combo.pack(fill=X, padx=5, pady=5)
    # Due Date
    twenty_days = [str(day) for day in range(1, 29)]
    thirty_days = [str(day) for day in range(1, 31)]
    thirtyone_days = [str(day) for day in range(1, 32)]

    months = [
        "January", "February",
        "March", "April",
        "May", "June",
        "July", "August",
        "September", "October",
        "November", "December"
    ]

    payment_make_payment_due_date_frame = Frame(payment_make_payment_frame)
    payment_make_payment_date_label = Label(payment_make_payment_due_date_frame, 
                            text="Date:",
                            font=("", 13, "bold"))

    payment_make_payment_date_day_var = StringVar()
    payment_make_payment_date_day_entry = ttk.Combobox(payment_make_payment_due_date_frame, 
                                        textvariable=payment_make_payment_date_day_var,
                                        values=thirtyone_days, font=("", 13, "bold"),
                                        width=4) 
    payment_make_payment_date_day_entry.insert(0, "Day")
    payment_make_payment_date_day_entry["state"] = "readonly"

    payment_make_payment_date_month_var = StringVar()
    payment_make_payment_date_month_entry = ttk.Combobox(payment_make_payment_due_date_frame,
                                        textvariable=payment_make_payment_date_month_var,
                                        values=months, font=("", 13, "bold"),
                                        width=10) 
    payment_make_payment_date_month_entry.insert(0, "Month")
    payment_make_payment_date_month_entry["state"] = "readonly"

    year = [str(y) for y in range(2020, 2101)]
    payment_make_payment_date_year_var = StringVar()
    payment_make_payment_date_year_entry = ttk.Combobox(payment_make_payment_due_date_frame, 
                                        textvariable=payment_make_payment_date_year_var,
                                        values=year, font=("", 13, "bold"), width=10) 
    payment_make_payment_date_year_entry.insert(0, "Year")
    payment_make_payment_date_year_entry["state"] = "readonly"

    payment_make_payment_due_date_frame.pack(fill=X, padx=5)
    payment_make_payment_date_label.pack(side=LEFT, anchor=W, padx=(5, 20))
    payment_make_payment_date_day_entry.pack(side=LEFT, anchor=W)
    payment_make_payment_date_month_entry.pack(side=LEFT, anchor=W, padx=5)
    payment_make_payment_date_year_entry.pack(side=LEFT, anchor=W)

    payment_make_pay_btn = Button(payment_make_payment_frame, text="MAKE PAYMENT", 
                            relief='flat', font=("", 13, "bold"), bg='#217346', 
                            command=lambda:verify(management=manage_payments, action="making_payment", unit_to_update=payment_make_payment_unit_entry.get().strip(),
                                                  amount=payment_make_payment_amount_entry.get().strip(), pay_name=payment_make_payment_name_entry.get().strip(), 
                                                  pay_for=payment_for_what_combo.get().strip(), date_day=payment_make_payment_date_day_var.get().strip(), 
                                                  date_month=payment_make_payment_date_month_var.get().strip(), date_year=payment_make_payment_date_year_var.get().strip(), 
                                                  function_for_action=make_payment))
    payment_make_pay_btn.pack(fill=X, padx=5, pady=5)

    set_ndd_entries_frame = ttk.LabelFrame(payment_entries_frame,
                                    text="Set New Due")

    set_ndd_unit_num = ttk.Entry(set_ndd_entries_frame, font=("", 13, "bold"))
    set_ndd_unit_num.insert(0, "Unit #")
    def ndd_unit_i(event):
        if set_ndd_unit_num.get().strip() == "" or \
            set_ndd_unit_num.get().strip() == "Unit #":
            set_ndd_unit_num.delete(0, END)
    def ndd_unit_o(event):
        if set_ndd_unit_num.get().strip() == "":
            set_ndd_unit_num.insert(0, "Unit #")
    set_ndd_unit_num.bind("<FocusIn>", ndd_unit_i)
    set_ndd_unit_num.bind("<FocusOut>", ndd_unit_o)

    set_ndd_due_date_frame = Frame(set_ndd_entries_frame)
    set_ndd_date_label = Label(set_ndd_due_date_frame, 
                            text="Date:",
                            font=("", 13, "bold"))

    set_nnd_date_day_var = StringVar()
    set_nnd_date_day_entry = ttk.Combobox(set_ndd_due_date_frame, 
                                        textvariable=set_nnd_date_day_var,
                                        values=thirtyone_days, font=("", 13, "bold"),
                                        width=4) 
    set_nnd_date_day_entry.insert(0, "Day")
    set_nnd_date_day_entry["state"] = "readonly"

    set_nnd_date_month_var = StringVar()
    set_nnd_date_month_entry = ttk.Combobox(set_ndd_due_date_frame, 
                                        textvariable=set_nnd_date_month_var,
                                        values=months, font=("", 13, "bold"),
                                        width=10) 
    set_nnd_date_month_entry.insert(0, "Month")
    set_nnd_date_month_entry["state"] = "readonly"

    year = [str(y) for y in range(2021, 2101)]
    set_nnd_date_year_var = StringVar()
    set_nnd_date_year_entry = ttk.Combobox(set_ndd_due_date_frame, 
                                        textvariable=set_nnd_date_year_var,
                                        values=year, font=("", 13, "bold"),
                                        width=10) 
    set_nnd_date_year_entry.insert(0, "Year")
    set_nnd_date_year_entry["state"] = "readonly"

    set_nnd_btn = Button(set_ndd_entries_frame, text="SET NEW DUE", font=("", 13, "bold"), bg='#217346', relief='flat', 
                         command=lambda:verify(management=manage_payments, action="updating_due", 
                                               function_for_action=payment_set_ndd, 
                                               unit_to_update=set_ndd_unit_num.get().strip(),
                                               date_day=set_nnd_date_day_var.get().strip(), 
                                               date_month=set_nnd_date_month_var.get().strip(), 
                                               date_year=set_nnd_date_year_var.get().strip()))

    set_ndd_entries_frame.pack(fill=X)
    set_ndd_unit_num.pack(fill=X, padx=5, pady=5)

    set_ndd_due_date_frame.pack(fill=X)
    set_ndd_date_label.pack(side=LEFT, anchor=W, padx=(10, 20))
    set_nnd_date_day_entry.pack(side=LEFT, anchor=W, pady=(0, 5))
    set_nnd_date_month_entry.pack(side=LEFT, anchor=W, padx=5, pady=(0, 5))
    set_nnd_date_year_entry.pack(side=LEFT, anchor=W, pady=(0, 5))
    set_nnd_btn.pack(fill=X, padx=5, pady=(0, 5))

    payment_apart_man_btn = ttk.Button(payment_right_frame, text="MANAGE APARTMENTS",
                                command=payment_apart)
    payment_tenant_btn = ttk.Button(payment_right_frame, text="MANAGE TENANTS",
                            command=payment_tenant)
    payment_exit_btn = ttk.Button(payment_right_frame, text="MAIN PAGE",
                                command=payment_exit)
    payment_apart_man_btn.pack(side=LEFT, pady=(5, 1), fill=X, expand=True, ipady=3)
    payment_tenant_btn.pack(side=LEFT, pady=(5, 1), fill=X, expand=True, padx=5, ipady=3)
    payment_exit_btn.pack(side=LEFT, pady=(5, 1), fill=X, expand=True,ipadx=35, ipady=3)

    excel_items = list(payment_main_sheet.values)
    payment_tree = ttk.Treeview(payment_tree_frame, show='headings', columns=excel_items[0])
    payment_scroll_bar_y = ttk.Scrollbar(payment_tree_frame, orient='vertical', command=payment_tree.yview)
    payment_scroll_bar_x = ttk.Scrollbar(payment_tree_frame, orient='horizontal', command=payment_tree.xview)
    payment_tree.configure(yscrollcommand=payment_scroll_bar_y.set,xscrollcommand=payment_scroll_bar_x.set)
    for heading in excel_items[0]:
        if heading == "UNIT":
            payment_tree.column(heading, anchor=CENTER, width=40)
            payment_tree.heading(heading, text=heading)
        elif heading.strip() == "STATUS":
            payment_tree.column(heading, anchor=CENTER, width=80)
            payment_tree.heading(heading, text=heading)
        elif heading.strip() == "TENANT":
            payment_tree.column(heading, anchor=CENTER, width=180)
            payment_tree.heading(heading, text=heading)
        elif heading.strip() == "DUE":
            payment_tree.column(heading, anchor=CENTER, width=180)
            payment_tree.heading(heading, text=heading)
        else:
            payment_tree.column(heading, anchor=CENTER, width=120)
            payment_tree.heading(heading, text=heading)
    for value in excel_items[1:]:
        payment_tree.insert('', END, values=value)
    payment_scroll_bar_y.pack(side=RIGHT, fill=Y, pady=(8, 0))
    payment_tree.pack(expand=TRUE, fill=Y, pady=(8, 0))

    manage_payments.mainloop()


def verify(management=None, action=None, unit_to_delete=0,
           unit_to_update=0, function_for_action=None, 
           amount=0, pay_name="Name", pay_for="PAYMENT FOR",
           date_day="Day", date_month="Month", date_year="Year",
           rent_stats="", utili_stat=""):
    is_continue = False
    actions_list = [
        "deleting_apartment",
        "deleting_tenant",
        "making_payment",
        "updating_apartment",
        "updating_tanant",
        "updating_due",
        "deleting_payment_history",
        ]
    

    def is_unit_exist(unti_n):
        unit_exist = False
        for every_due in range(2, apart_sheet.max_row+1):
            if unti_n == apart_sheet["A"+str(every_due)].value:
                unit_exist = True
                break
        return unit_exist
    

    def is_unit_has_tenant(unit_n):
        unit_has_tenant = False
        for every_tenant in range(2, tenant_main_sheet.max_row+1):
            if unit_n == tenant_main_sheet["A"+str(every_tenant)].value:
                unit_has_tenant = True
                break
        return unit_has_tenant
    
    # deleting_apartment
    if action == actions_list[0]:
        if unit_to_delete == "Unit #" or unit_to_delete == "":
            messagebox.showerror('ERROR', "Must provide a Unit Number.")
        else:
            has_tenant = is_unit_has_tenant(unit_to_delete)
            if has_tenant:
                messagebox.showerror("ERROR", f"Can't perform action, Unit \"{unit_to_delete}\" is occupied.")
            else:
                search_unit_data_found = is_unit_exist(unit_to_delete)
                if search_unit_data_found:
                    is_continue = True
                else:
                    messagebox.showerror("ERROR", f"There is no existing data for Unit \"{unit_to_delete}\"")
    # deleting_tenant
    elif action == actions_list[1]:
        if unit_to_delete == "Unit #" or unit_to_delete == "":
            messagebox.showerror('ERROR', "Must provide a Unit Number.")
        else:
            unit_found = is_unit_exist(unit_to_delete)
            if unit_found:
                it_has_tenant = is_unit_has_tenant(unit_to_delete)
                if it_has_tenant:
                    has_due = list(payment_main_sheet["A"])
                    tenant_that_has_due = []
                    is_tenant_h_d = False 
                    u_r = 0
                    for every_due_r in range(2, payment_main_sheet.max_row+1):
                        if unit_to_delete == payment_main_sheet["A"+str(every_due_r)].value:
                            u_r = every_due_r
                            break
                    for e_d in has_due:
                        tenant_that_has_due.append(e_d.value)
                    if unit_to_delete in tenant_that_has_due:
                        if payment_main_sheet["G"+str(u_r)].value.lower() != "none":
                            is_tenant_h_d = True
                    if is_tenant_h_d:
                        messagebox.showerror('ERROR', f'Tenant in Unit "{unit_to_delete}" has remaining balance. \nThis must be settled first.')
                    else:
                        is_continue = True
                else:
                    messagebox.showerror("ERROR", f"Unit \"{unit_to_delete}\" vacant.")
            else:
                messagebox.showerror("ERROR", f"There is no existing data for Unit \"{unit_to_delete}\".")
    # making_payment
    elif action == actions_list[2]:
        # Check if there is unit number inputed
        if unit_to_update != "Unit #" and unit_to_update != "":
            has_unit_number = True
        elif unit_to_update == "Unit #":
            has_unit_number = False
        # Check if at least one entry aside unit number has input
        atleast_one = False
        if amount == "Amount" and pay_name == "Name" \
            and pay_for == "PAYMENT FOR" and date_day == "Day" \
            and date_month == "Month" and date_year == "Year":
            atleast_one = False
        elif amount != "Amount" or pay_name != "Name" \
            or pay_for != "PAYMENT FOR" or date_day != "Day" \
            or date_month != "Month" or date_year != "Year":
            atleast_one = True

        unit_error_unit = False
        if unit_to_update == "":
            unit_error_unit = True
        elif has_unit_number == False and atleast_one == True:
            unit_error_unit = True
        elif has_unit_number == True and atleast_one == False:
            check_unit = True
        elif has_unit_number == False and atleast_one == False:
            check_unit = False
        elif has_unit_number == True and atleast_one == True:
            check_unit = True

        if unit_error_unit:
            messagebox.showerror('ERROR', "Must provide a Unit Number.")
        elif check_unit:
            unit_exist = is_unit_exist(unit_to_update)
            if unit_exist:
                unit_exist = is_unit_exist(unit_to_update)
                if unit_exist:
                    is_unit_f = False
                    for every_d in range(2, payment_main_sheet.max_row+1):
                        if unit_to_update == payment_main_sheet[f"A{every_d}"].value:
                            is_unit_f = True
                            break
                    if is_unit_f:
                        if unit_to_update != "" and unit_to_update != "Unit #" \
                            and amount != "" and amount != "Amount" \
                            and pay_name != "" and pay_name != "Name" \
                            and pay_for != "" and pay_for != "PAYMENT FOR" \
                            and date_day != "" and date_day != "Day" \
                            and date_month != "" and date_month != "Month" \
                            and date_year != "" and date_year != "Year":
                            is_continue = True
                        else:
                            messagebox.showerror('ERROR', "Must fill all fields to continue.")
                    else:
                        messagebox.showerror("ERROR", f"Unit {unit_to_update} has no set due date.")
                else:
                    messagebox.showerror("ERROR", f"Unit {unit_to_update} is vacant.")
            else:
                messagebox.showerror("ERROR", f"There is no existing data for Unit \"{unit_to_update}\".")
        else:
            pass
    # updating_apartment
    elif action == actions_list[3]:
        does_exist = is_unit_exist(unit_to_update)
        if unit_to_update != "Unit #" and unit_to_update != "":
            if does_exist:
                is_continue = True
            else:
                messagebox.showerror("ERROR", f"There is no existing data for Unit \"{unit_to_update}\".")
        else:
            messagebox.showerror('ERROR', "Must provide a Unit Number.")
    # updating_tanant
    elif action == actions_list[4]:
        u_exist = is_unit_exist(unit_to_update)
        if unit_to_update != "" and unit_to_update != "Unit #":
            if u_exist:
                u_has_tenant = is_unit_has_tenant(unit_to_update)
                if u_has_tenant:
                    is_continue = True
                else:
                    messagebox.showerror("ERROR", f"Unit \"{unit_to_update}\" is vacant.")
            else:
                messagebox.showerror("ERROR", f"There is no existing data for Unit \"{unit_to_update}\".")
        else:
            messagebox.showerror('ERROR', "Must provide a Unit Number.")
    # updating_due
    elif action == actions_list[5]:
        if unit_to_update != "Unit #" and unit_to_update != "":
            has_unit_number = True
        elif unit_to_update == "Unit #":
            has_unit_number = False

        atleast_one = False
        if date_day == "Day" and date_month == "Month"\
            and date_year == "Year":
            atleast_one = False
        elif date_day != "Day" and date_month != "Month"\
            and date_year != "Year":
            atleast_one = True

        unit_error_unit = False
        if unit_to_update == "":
            unit_error_unit = True
        elif has_unit_number == False and atleast_one == True:
            unit_error_unit = True
        elif has_unit_number == True and atleast_one == False:
            check_unit = True
        elif has_unit_number == False and atleast_one == False:
            check_unit = False
        elif has_unit_number == True and atleast_one == True:
            check_unit = True

        if unit_error_unit:
            messagebox.showerror('ERROR', "Must provide a Unit Number.")
        elif check_unit:
            d_exist = (unit_to_update)
            if d_exist:
                t_exist = is_unit_has_tenant(unit_to_update)
                if t_exist:
                    if unit_to_update != "Unit #" and unit_to_update != ""\
                        and date_day != "Day" and date_day != ""\
                        and date_month != "Month" and date_month != ""\
                        and date_year != "Year" and date_year != "":
                        is_continue = True
                    else:
                        messagebox.showerror('ERROR', "Must fill all fields to continue.")
                else:
                    messagebox.showerror("ERROR", f"Unit {unit_to_update} is vacant.")
            else:
                messagebox.showerror("ERROR", f"There is no existing data for Unit \"{unit_to_update}\".")
        else:
            pass
    # deleting_payment_history
    elif action == actions_list[6]:
        if unit_to_delete != "Unit #" and unit_to_delete != "":
            row = 0
            found = False
            for i in payment_record_sheet["A"]:
                row += 1
                if unit_to_delete == i.value:
                    found = True
                    break
            if found:
                history_list = list(payment_record_sheet[row])
                history_list_value = []
                for e_h in history_list:
                    if e_h.value != None:
                        history_list_value.append(e_h.value)
                history_list_value_length = len(history_list_value)
                if history_list_value_length < 2:
                    messagebox.showerror("ERROR", f"Can't find payment record for \"{unit_to_delete}\"")
                else:
                    is_continue = True
            else:
                messagebox.showerror("ERROR", f"Can't find payment record for Unit \"{unit_to_delete}\"")
        else:
            messagebox.showerror('ERROR', "Must provide a Unit Number.")

    def to_verify_btn_func():
        code = '1'
        is_verify = False
        user_enrty = veri_code_entry.get().strip()
        if user_enrty != "Enter code" and user_enrty != "":
            if user_enrty == code:
                is_verify = True
            else:
                veri_code_entry.delete(0, END)
                messagebox.showinfo("Incorrect", "Incorrect code, try again.")
        else:
            messagebox.showerror("ERROR", "Must enter code to continue action.")
        if is_verify:
            if action == actions_list[0]:
                function_for_action(unit_to_delete, to_verify)
            elif action == actions_list[1]:
                function_for_action(unit_to_delete, to_verify)
            elif action == actions_list[2]:
                function_for_action(unit_to_update, to_verify, date_day, date_month, date_year, pay_for, amount, pay_name)
            elif action == actions_list[3]:
                function_for_action(unit_to_update, to_verify)
            elif action == actions_list[4]:
                function_for_action(unit_to_update, to_verify)
            elif action == actions_list[5]:
                function_for_action(unit_to_update, to_verify, date_day, date_month, date_year)
            elif action == actions_list[6]:
                function_for_action(unit_to_delete, to_verify)
    

    def to_cancel_btn_func():
        to_verify.destroy()
        if action == actions_list[0]:
            apartment()
        elif action == actions_list[1]:
            tenant()
        elif action == actions_list[2]:
            payment()
        elif action == actions_list[3]:
            apartment()
        elif action == actions_list[4]:
            tenant()
        elif action == actions_list[5]:
            payment()
        elif action == actions_list[6]:
            payment()
    if is_continue:
        management.destroy()
        to_verify = tk.Tk()
        to_verify.title("Verifying")
        to_verify.geometry("340x340")
        to_verify.resizable(0, 0)
        # Theme
        style = ttk.Style(to_verify)
        to_verify.tk.call("source", 'theme_forest-dark.tcl')
        style.theme_use("forest-dark")

        veri_main_frame = Frame(to_verify)
        veri_main_frame.pack(expand=TRUE, fill=BOTH, padx=20, pady=20)
        veri_top_frame = Frame(veri_main_frame)
        veri_top_frame.pack(fill=X)
        veri_label = Label(veri_top_frame, text="VERIFYING", font=("Arial Black", 35), fg="#217346")
        veri_label.pack(pady=(20, 20))

        veri_entry_frame = Frame(veri_main_frame)
        veri_entry_frame.pack(fill=BOTH)
        veri_code_label_frame = Frame(veri_entry_frame)
        veri_code_label_frame.pack(fill=X)
        veri_code_label = ttk.Label(veri_code_label_frame, text="ENTER CODE:", font=("", 13, 'bold'))
        veri_code_label.pack(anchor=W, pady=(0, 10))
        veri_code_entry = ttk.Entry(veri_entry_frame, font=("", 11))
        veri_code_entry.insert(0, "CODE")
        def veri_code_entry_i(event):
            if veri_code_entry.get().strip() == "" or \
                veri_code_entry.get().strip() == "CODE":
                veri_code_entry.delete(0, END)
                veri_code_entry['show'] = "•"
        def veri_code_entry_o(event):
            if veri_code_entry.get().strip() == "":
                veri_code_entry.insert(0, "CODE")
                veri_code_entry['show'] = ""
        veri_code_entry.bind("<FocusIn>", veri_code_entry_i)
        veri_code_entry.bind("<FocusOut>", veri_code_entry_o)
        veri_code_entry.pack(fill=X, pady=(0, 20), ipady=5)

        verify_btn = Button(veri_main_frame, text="CONTINUE", relief='flat', font=("Arial", 10), bg='#217346', command=to_verify_btn_func)
        cancel_btn = ttk.Button(veri_main_frame, text="CANCEL", command=to_cancel_btn_func)
        verify_btn.pack(fill=X, pady=(0, 7), ipady=4)
        cancel_btn.pack(fill=X, ipady=2)

        to_verify.mainloop()


log_in_interface()

